#!/usr/bin/env python
# -*- coding: UTF-8 -*-
"""
@Project ：wallpaper
@File    ：view.py
@Author  ：Liang
@Date    ：2026/4/28
@description : 页面TDK管理
"""
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from rest_framework import serializers
from rest_framework.decorators import action

from models.models import PageTDK, SiteConfig
from tool.base_views import BaseViewSet
from tool.permissions import IsAdmin
from tool.utils import ApiResponse, CustomPagination


class PageTDKSerializer(serializers.ModelSerializer):
    """页面TDK序列化器"""
    page_type_display = serializers.CharField(source='get_page_type_display', read_only=True)
    url_content = serializers.SerializerMethodField(help_text="关联的URL地址")

    class Meta:
        model = PageTDK
        fields = [
            'id', 'page_type', 'page_type_display', 'title', 'description',
            'keywords', 'url', 'url_content', 'applied_count',
            'is_template', 'is_active', 'created_at', 'updated_at'
        ]
        read_only_fields = ['id', 'created_at', 'updated_at']

    def get_url_content(self, obj):
        """获取关联的URL地址"""
        if obj.url:
            return obj.url.content
        return None


@extend_schema(tags=["页面TDK管理"])
@extend_schema_view(
    list=extend_schema(
        summary="获取页面TDK列表",
        description="支持按页面类型、是否模板、是否启用筛选，"
                    "page_type:category,tag,detail,search,article,custom",
        parameters=[
            OpenApiParameter(name="currentPage", type=int, required=False, description="当前页码"),
            OpenApiParameter(name="pageSize", type=int, required=False, description="每页数量"),
            OpenApiParameter(name="is_template", type=str, required=False, description="是否模板"),
            OpenApiParameter(name="url", type=str, required=False, description="是否模板"),
        ],
    ),
    retrieve=extend_schema(summary="获取页面TDK详情"),
    create=extend_schema(summary="创建页面TDK"),
    update=extend_schema(summary="更新页面TDK"),
    partial_update=extend_schema(summary="部分更新页面TDK"),
    destroy=extend_schema(summary="删除页面TDK"),
    detect_duplicate_titles=extend_schema(
        summary="检测重复标题",
        description="检测所有页面TDK中重复的Title，返回重复标题、涉及页面URL列表、重复次数。只检测已启用的TDK记录。",
    ),
    export_tdk_report=extend_schema(
        summary="导出TDK报告",
        description="导出所有页面TDK配置为CSV或Excel格式，包含：页面URL、页面类型、Title、Description、Keywords、应用次数等真实数据",
        parameters=[
            OpenApiParameter(name="format", type=str, required=False, description="导出格式：csv或excel，默认csv"),
            OpenApiParameter(name="is_active", type=str, required=False, description="是否只导出启用的：true/false，默认全部"),
        ],
    ),
    import_tdk_report=extend_schema(
        summary="导入TDK报告",
        description="从CSV/Excel文件导入TDK配置，以URL为主键进行匹配更新。如果URL已存在则更新，不存在则创建。支持批量导入。",
    ),
)
class PageTDKViewSet(BaseViewSet):
    """
    页面TDK管理 ViewSet
    管理页面的Title、Description、Keywords配置
    """
    queryset = PageTDK.objects.all()
    serializer_class = PageTDKSerializer
    pagination_class = CustomPagination
    # permission_classes = [IsAdmin]

    def get_queryset(self):
        queryset = super().get_queryset()
        # 按页面类型筛选
        page_type = self.request.query_params.get('page_type')
        if page_type:
            queryset = queryset.filter(page_type=page_type)
        # 按是否模板筛选
        is_template = self.request.query_params.get('is_template')
        if is_template is not None:
            queryset = queryset.filter(is_template=is_template.lower() == 'true')
        # 按是否启用筛选
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        url = self.request.query_params.get('url', '').strip()
        if url:
            queryset = queryset.filter(url__content__icontains=url)

        return queryset.order_by('-updated_at')

    def list(self, request, *args, **kwargs):
        """获取页面TDK列表"""
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return ApiResponse(data=serializer.data, message="列表获取成功")

    def retrieve(self, request, *args, **kwargs):
        """获取页面TDK详情"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return ApiResponse(data=serializer.data, message="获取成功")

    def create(self, request, *args, **kwargs):
        """创建页面TDK"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        result_serializer = self.get_serializer(instance)
        return ApiResponse(data=result_serializer.data, message="创建成功", code=201)

    def update(self, request, *args, **kwargs):
        """更新页面TDK"""
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        result_serializer = self.get_serializer(instance)
        return ApiResponse(data=result_serializer.data, message="更新成功")

    def partial_update(self, request, *args, **kwargs):
        """部分更新页面TDK"""
        return self.update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """删除页面TDK"""
        instance = self.get_object()
        instance.delete()
        return ApiResponse(message="删除成功")
    
    @extend_schema(
        summary="检测重复标题",
        description="检测所有页面TDK中重复的Title，返回重复标题、涉及页面URL列表、重复次数",
    )
    @action(detail=False, methods=['get'], name='检测重复标题')
    def detect_duplicate_titles(self, request):
        """
        检测重复标题
        返回格式：[
            {
                "title": "重复的标题内容",
                "duplicate_count": 3,
                "urls": [
                    {"url": "https://example.com/page1", "page_type": "detail"},
                    {"url": "https://example.com/page2", "page_type": "article"}
                ]
            }
        ]
        """
        from django.db.models import Count
        
        # 查询所有启用的TDK记录
        queryset = PageTDK.objects.filter(is_active=True)
        
        # 按标题分组，统计每个标题出现的次数
        title_counts = queryset.values('title').annotate(
            count=Count('id')
        ).filter(count__gt=1)  # 只保留重复的标题
        
        duplicate_results = []
        
        for item in title_counts:
            title = item['title']
            if not title:  # 跳过空标题
                continue
            
            # 获取使用该标题的所有页面
            pages = queryset.filter(title=title).values('url__content', 'page_type')
            
            urls = []
            for page in pages:
                if page['url__content']:  # 只包含有URL的记录
                    urls.append({
                        'url': page['url__content'],
                        'page_type': page['page_type']
                    })
            
            if urls:  # 只返回有URL的重复项
                duplicate_results.append({
                    'title': title,
                    'duplicate_count': len(urls),
                    'urls': urls
                })
        
        # 按重复次数降序排列
        duplicate_results.sort(key=lambda x: x['duplicate_count'], reverse=True)
        
        return ApiResponse(
            data={
                'total_duplicates': len(duplicate_results),
                'duplicates': duplicate_results
            },
            message=f"检测到{len(duplicate_results)}个重复标题"
        )
    
    @extend_schema(
        summary="导出TDK报告",
        description="生成TDK报告并返回下载链接，支持CSV和Excel格式",
        parameters=[
            OpenApiParameter(name="export_format", type=str, required=False, description="导出格式：csv或excel，默认csv"),
            OpenApiParameter(name="is_active", type=str, required=False, description="是否只导出启用的：true/false，默认全部"),
        ],
    )
    @action(detail=False, methods=['get'], url_path='export-tdk-report', name='导出TDK报告')
    def export_tdk_report(self, request):
        """
        导出TDK报告
        返回下载链接
        
        示例：
        GET /api/seo/tdk/export-tdk-report/?export_format=csv&is_active=true
        GET /api/seo/tdk/export-tdk-report/?export_format=excel
        """
        # 获取参数
        export_format = request.query_params.get('export_format', 'csv').lower().strip()
        is_active = request.query_params.get('is_active')
        
        # 验证参数
        if export_format not in ['csv', 'excel']:
            return ApiResponse(code=400, message="不支持的导出格式，请使用csv或excel")
        
        # 构建查询集
        queryset = PageTDK.objects.all().order_by('-updated_at')
        
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        # 统计数量
        count = queryset.count()
        
        # 生成下载URL（使用当前请求的host）
        host = request.get_host()
        scheme = request.scheme
        download_url = f"{scheme}://{host}/api/seo/tdk/export-tdk-download/?export_format={export_format}"
        
        if is_active is not None:
            download_url += f"&is_active={is_active}"
        
        return ApiResponse(
            data={
                'download_url': download_url,
                'format': export_format,
                'count': count
            },
            message=f"成功导出{count}条数据"
        )
    
    @action(detail=False, methods=['get'], url_path='export-tdk-download', name='下载TDK文件', permission_classes=[])
    def export_tdk_download(self, request):
        """
        下载TDK报告文件（无需认证）
        直接返回文件流
        
        示例：
        GET /api/seo/tdk/export-tdk-download/?export_format=csv
        GET /api/seo/tdk/export-tdk-download/?export_format=excel
        """
        import csv
        import io
        from django.http import HttpResponse
        
        # 获取参数
        export_format = request.query_params.get('export_format', 'csv').lower().strip()
        is_active = request.query_params.get('is_active')
        
        # 构建查询集
        queryset = PageTDK.objects.all().order_by('-updated_at')
        
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        # 获取所有数据
        tdk_list = list(queryset.values(
            'id', 'page_type', 'title', 'description', 'keywords',
            'url__content', 'applied_count', 'is_template', 'is_active',
            'created_at', 'updated_at'
        ))
        
        if export_format == 'csv':
            # 生成CSV文件
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 写入表头
            writer.writerow([
                'ID', '页面类型', '标题(Title)', '描述(Description)',
                '关键词(Keywords)', '页面URL', '应用次数', '是否模板',
                '是否启用', '创建时间', '更新时间'
            ])
            
            # 写入数据
            for item in tdk_list:
                writer.writerow([
                    item['id'],
                    item['page_type'],
                    item['title'] or '',
                    item['description'] or '',
                    item['keywords'] or '',
                    item['url__content'] or '',
                    item['applied_count'],
                    '是' if item['is_template'] else '否',
                    '是' if item['is_active'] else '否',
                    item['created_at'].strftime('%Y-%m-%d %H:%M:%S') if item['created_at'] else '',
                    item['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if item['updated_at'] else ''
                ])
            
            # 设置响应
            response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="tdk_report.csv"'
            
        elif export_format == 'excel':
            # 生成Excel文件（需要openpyxl库）
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "TDK报告"
                
                # 设置表头样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 写入表头
                headers = [
                    'ID', '页面类型', '标题(Title)', '描述(Description)',
                    '关键词(Keywords)', '页面URL', '应用次数', '是否模板',
                    '是否启用', '创建时间', '更新时间'
                ]
                ws.append(headers)
                
                # 设置表头样式
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 写入数据
                for item in tdk_list:
                    ws.append([
                        item['id'],
                        dict(PageTDK.PAGE_TYPE_CHOICES).get(item['page_type'], item['page_type']),
                        item['title'] or '',
                        item['description'] or '',
                        item['keywords'] or '',
                        item['url__content'] or '',
                        item['applied_count'],
                        '是' if item['is_template'] else '否',
                        '是' if item['is_active'] else '否',
                        item['created_at'].strftime('%Y-%m-%d %H:%M:%S') if item['created_at'] else '',
                        item['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if item['updated_at'] else ''
                    ])
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 保存为字节流
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # 设置响应
                response = HttpResponse(
                    output.read(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = 'attachment; filename="tdk_report.xlsx"'
                
            except ImportError:
                return HttpResponse(
                    "Excel导出需要安装openpyxl库：pip install openpyxl",
                    status=400
                )
        else:
            return HttpResponse("不支持的导出格式", status=400)
        
        return response
    
    @extend_schema(
        summary="导入TDK报告",
        description="从CSV/Excel文件导入TDK配置，以URL为主键进行匹配更新。如果URL已存在则更新，不存在则创建。支持批量导入。",
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {"type": "string", "format": "binary", "description": "CSV或Excel文件 (.csv, .xlsx, .xls)"}
                },
                "required": ["file"]
            }
        },
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "导入完成！成功50条（更新30条，新建20条）"},
                    "data": {
                        "type": "object",
                        "properties": {
                            "success_count": {"type": "integer", "example": 50},
                            "update_count": {"type": "integer", "example": 30},
                            "create_count": {"type": "integer", "example": 20},
                            "error_count": {"type": "integer", "example": 0},
                            "errors": {"type": "array", "items": {"type": "string"}}
                        }
                    }
                }
            },
            400: {"description": "文件格式错误或数据异常"},
            500: {"description": "服务器内部错误"}
        }
    )
    @action(detail=False, methods=['post'], name='导入TDK报告')
    def import_tdk_report(self, request):
        """
        导入TDK报告
        支持CSV和Excel格式，以URL为主键匹配
        """
        import csv
        import io
        
        # 检查是否有文件上传
        if 'file' not in request.FILES:
            return ApiResponse(code=400, message="请上传文件")
        
        uploaded_file = request.FILES['file']
        file_name = uploaded_file.name.lower()
        
        # 读取文件内容
        try:
            if file_name.endswith('.csv'):
                # 处理CSV文件
                content = uploaded_file.read().decode('utf-8-sig')  # 处理BOM
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)
            elif file_name.endswith(('.xlsx', '.xls')):
                # 处理Excel文件
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filename=io.BytesIO(uploaded_file.read()))
                    ws = wb.active
                    
                    # 获取表头
                    headers = [cell.value for cell in ws[1]]
                    
                    # 转换为字典列表
                    rows = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if any(row):  # 跳过空行
                            row_dict = {}
                            for header, value in zip(headers, row):
                                # 将None转换为空字符串，保证后续处理安全
                                row_dict[header] = value if value is not None else ''
                            rows.append(row_dict)
                except ImportError:
                    return ApiResponse(
                        code=400,
                        message="Excel导入需要安装openpyxl库：pip install openpyxl"
                    )
            else:
                return ApiResponse(code=400, message="只支持CSV或Excel文件格式")
            
            if not rows:
                return ApiResponse(code=400, message="文件中没有数据")
            
            # 处理数据
            success_count = 0
            update_count = 0
            create_count = 0
            error_count = 0
            errors = []
            
            for index, row in enumerate(rows, start=2):  # 从第2行开始（第1行是表头）
                try:
                    # 获取URL（必需字段）
                    url_content = row.get('页面URL', '').strip()
                    if not url_content:
                        errors.append(f"第{index}行：缺少页面URL")
                        error_count += 1
                        continue
                    
                    # 查找是否已存在该URL的记录
                    existing_tdk = PageTDK.objects.filter(url__content=url_content).first()
                    
                    # 准备数据
                    tdk_data = {
                        'title': row.get('标题(Title)', '').strip(),
                        'description': row.get('描述(Description)', '').strip(),
                        'keywords': row.get('关键词(Keywords)', '').strip(),
                        'page_type': row.get('页面类型', 'custom').strip(),
                    }
                    
                    # 验证page_type
                    valid_page_types = [choice[0] for choice in PageTDK.PAGE_TYPE_CHOICES]
                    if tdk_data['page_type'] not in valid_page_types:
                        tdk_data['page_type'] = 'custom'  # 默认值
                    
                    # 处理布尔值
                    is_template_val = row.get('是否模板', '否').strip()
                    tdk_data['is_template'] = is_template_val in ['是', 'true', 'True', '1']
                    
                    is_active_val = row.get('是否启用', '是').strip()
                    tdk_data['is_active'] = is_active_val in ['是', 'true', 'True', '1']
                    
                    if existing_tdk:
                        # 更新现有记录
                        for key, value in tdk_data.items():
                            setattr(existing_tdk, key, value)
                        existing_tdk.save()
                        update_count += 1
                    else:
                        # 创建新记录，需要先找到或创建SiteConfig对象
                        url_obj, _ = SiteConfig.objects.get_or_create(
                            content=url_content,
                            defaults={
                                'config_type': 'sitemap_url',
                                'is_active': True
                            }
                        )
                        
                        PageTDK.objects.create(
                            url=url_obj,
                            **tdk_data
                        )
                        create_count += 1
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"第{index}行：{str(e)}")
                    error_count += 1
            
            # 构建响应消息
            message = f"导入完成！成功{success_count}条（更新{update_count}条，新建{create_count}条）"
            if error_count > 0:
                message += f"，失败{error_count}条"
            
            result_data = {
                'success_count': success_count,
                'update_count': update_count,
                'create_count': create_count,
                'error_count': error_count,
            }
            
            if errors:
                result_data['errors'] = errors[:10]  # 只返回前10个错误
            
            return ApiResponse(data=result_data, message=message)
            
        except Exception as e:
            return ApiResponse(code=500, message=f"导入失败：{str(e)}")
