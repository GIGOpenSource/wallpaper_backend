"""
@Project ：wallpaper
@File    ：view.py
@Author  ：LiangHB
@Date    ：2026/4/14 17:13
@description : 壁纸相关视图逻辑
"""
import io
import json
import os
import uuid

from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import F
from django.db.models.functions import Greatest
from PIL import Image
from django.utils import timezone

from App.view.wallpapers.search_models.search_models import TAG_MAPPING
from tool.base_views import BaseViewSet
from tool.middleware import logger
from tool.permissions import IsCustomerTokenValid, IsOwnerOrAdmin, IsAdmin
from tool.token_tools import CustomTokenTool
from tool.uploader_data import bytes_from_uploaded_image, upload_image_to_cos
from tool.utils import CustomPagination, ApiResponse
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter, OpenApiRequest, OpenApiExample
from django.utils.translation import get_language, gettext as _, activate
import pandas as pd
from rest_framework.decorators import api_view, action
from rest_framework import serializers
from rest_framework.parsers import MultiPartParser, FormParser
from models.models import (
    Wallpapers,
    WallpaperTag,
    WallpaperCategory,
    NavigationTag,
    WallpaperLike,
    WallpaperCollection,
    CustomerWallpaperUpload,
    CustomerUser, RecommendStrategy, StrategyWallpaperRelation, SiteConfig,
)
from tool.operation_log import log_operation

def _parse_tag_ids(raw):
    if raw is None or raw == "":
        return []
    if isinstance(raw, list):
        out = []
        for x in raw:
            try:
                out.append(int(x))
            except (TypeError, ValueError):
                continue
        return out
    s = str(raw).strip()
    if not s:
        return []
    if s.startswith("["):
        try:
            arr = json.loads(s)
            out = []
            for x in arr:
                try:
                    out.append(int(x))
                except (TypeError, ValueError):
                    continue
            return out
        except json.JSONDecodeError:
            return []
    return [int(x) for x in s.split(",") if x.strip().isdigit()]


def _parse_tag_names(raw):
    if raw is None or raw == "":
        return []
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    s = str(raw).strip()
    if not s:
        return []
    if s.startswith("["):
        try:
            arr = json.loads(s)
            return [str(x).strip() for x in arr if str(x).strip()]
        except json.JSONDecodeError:
            return []
    return [x.strip() for x in s.split(",") if x.strip()]


def _person_wallpaper_cos_key(title: str, original_filename: str) -> tuple[str, str]:
    """COS 对象键 person_wallpaper/{title}.{ext}，扩展名随上传文件。"""
    _, ext = os.path.splitext(original_filename or "")
    ext = (ext or ".jpg").lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        ext = ".jpg"
    base = (title or "").strip()
    for ch in '<>:"|?*\\/\x00':
        base = base.replace(ch, "")
    base = base.strip(" .")[:180]
    if not base:
        base = uuid.uuid4().hex[:12]
    return f"person_wallpaper/{base}{ext}", ext.lstrip(".").lower()


def _image_meta_from_bytes(content: bytes):
    try:
        with Image.open(io.BytesIO(content)) as im:
            fmt = (im.format or "").lower() or None
            return im.width, im.height, fmt
    except Exception:
        return 0, 0, None


def _decrement_user_counters_on_wallpaper_delete(wallpaper_ids):
    """
    删除壁纸时，同步扣减相关用户的计数

    Args:
        wallpaper_ids: 壁纸ID列表（可以是单个ID或列表）

    Returns:
        dict: 统计信息 {'deleted_count': int, 'affected_users': int}
    """
    from django.db.models import Count
    # 确保是列表格式
    if not isinstance(wallpaper_ids, list):
        wallpaper_ids = [wallpaper_ids]
    if not wallpaper_ids:
        return {'deleted_count': 0, 'affected_users': 0}
    # 去重
    wallpaper_ids = list(set(wallpaper_ids))
    # 记录需要更新的计数器 {customer_id: {'upload': count, 'collect': count}}
    user_counters = {}
    # 1. 处理上传者计数
    uploads = CustomerWallpaperUpload.objects.filter(
        wallpaper_id__in=wallpaper_ids
    ).select_related('customer')
    for upload in uploads:
        cid = upload.customer_id
        if cid not in user_counters:
            user_counters[cid] = {'upload': 0, 'collect': 0}
        user_counters[cid]['upload'] += 1
    # 2. 处理收藏者计数（批量查询优化）
    collections = WallpaperCollection.objects.filter(wallpaper_id__in=wallpaper_ids)
    if collections.exists():
        collect_stats = collections.values('user_id').annotate(cnt=Count('user_id'))
        for stat in collect_stats:
            uid = stat['user_id']
            if uid not in user_counters:
                user_counters[uid] = {'upload': 0, 'collect': 0}
            user_counters[uid]['collect'] += stat['cnt']
    # 3. 批量更新用户计数（使用 Greatest 防止负数）
    affected_users = 0
    try:
        for cid, counts in user_counters.items():
            updated = CustomerUser.objects.filter(id=cid).update(
                upload_count=Greatest(F('upload_count') - counts['upload'], 0),
                collection_count=Greatest(F('collection_count') - counts['collect'], 0)
            )
            if updated:
                affected_users += 1
    except Exception as e:
        logger.error(f"删除壁纸后更新用户计数失败: {e}", exc_info=True)
    return {
        'deleted_count': len(wallpaper_ids),
        'affected_users': affected_users
    }


# ====================优化查询218start===============================
class WallpapersListSerializer(serializers.ModelSerializer):
    """壁纸列表序列化器（轻量级，只包含必要字段）"""
    # tags = serializers.SerializerMethodField()
    aspect_ratio = serializers.SerializerMethodField()
    # is_liked = serializers.SerializerMethodField()
    # is_collected = serializers.SerializerMethodField()

    class Meta:


        model = Wallpapers
        fields = [
            'id', 'name', 'url', 'thumb_url', 'width', 'height', 'image_format',
            'has_watermark', 'is_live', 'is_hd', 'hot_score', 'like_count',
            'collect_count', 'download_count', 'view_count', 'created_at',
            'aspect_ratio','audit_status'
        ]
        read_only_fields = fields

    def get_tags(self, obj):
        return [{'id': tag.id, 'name': tag.name} for tag in obj.tags.all()]

    def get_aspect_ratio(self, obj):
        if not obj.width or not obj.height:
            return None

        def gcd(a, b):
            while b:
                a, b = b, a % b
            return a

        common_divisor = gcd(obj.width, obj.height)
        return f"{obj.width // common_divisor}:{obj.height // common_divisor}"

    def get_is_liked(self, obj):
        liked_ids = self.context.get("liked_wallpaper_ids")
        return obj.id in liked_ids if liked_ids else False

    def get_is_collected(self, obj):
        collected_ids = self.context.get("collected_wallpaper_ids")
        return obj.id in collected_ids if collected_ids else False


class WallpapersAdminListSerializer(serializers.ModelSerializer):
    """壁纸列表序列化器（管理员使用，包含详细信息）"""
    tags = serializers.SerializerMethodField()
    category = serializers.SerializerMethodField()
    uploader = serializers.SerializerMethodField()
    aspect_ratio = serializers.SerializerMethodField()

    class Meta:
        model = Wallpapers
        fields = [
            'id', 'name', 'url', 'thumb_url', 'width', 'height', 'image_format',
            'source_url', 'description', 'has_watermark', 'category', 'tags',
            'is_live', 'is_hd', 'hot_score', 'like_count', 'collect_count',
            'download_count', 'view_count', 'created_at', 'aspect_ratio',
            'audit_status', 'uploader', 'audit_remark', 'audited_at','description'
        ]
        read_only_fields = fields

    def get_tags(self, obj):
        return [
            {
                'id': tag.id,
                'name': tag.name,
            }
            for tag in obj.tags.all()
        ]

    def get_category(self, obj):
        return [
            {
                'id': cat.id,
                'name': cat.name,
            }
            for cat in obj.category.all()
        ]

    def get_uploader(self, obj):
        """获取上传者信息"""
        try:
            upload_record = getattr(obj, 'customer_upload', None)
            if not upload_record:
                return None
            customer = getattr(upload_record, 'customer', None)
            if not customer:
                return None
            return {
                "id": customer.id,
                "email": customer.email,
                "nickname": customer.nickname,
                "gender": customer.gender,
                "avatar_url": customer.avatar_url,
                "badge": customer.badge,
                "points": customer.points,
                "level": customer.level,
                "upload_count": customer.upload_count,
                "collection_count": customer.collection_count,
                "last_login": customer.last_login.isoformat() if customer.last_login else None,
                "created_at": customer.created_at.isoformat() if customer.created_at else None,
            }
        except ObjectDoesNotExist:
            return None
    def get_aspect_ratio(self, obj):
        """计算宽高比，格式如 16:9"""
        if not obj.width or not obj.height:
            return None
        def gcd(a, b):
            while b:
                a, b = b, a % b
            return a
        common_divisor = gcd(obj.width, obj.height)
        width_ratio = obj.width // common_divisor
        height_ratio = obj.height // common_divisor
        return f"{width_ratio}:{height_ratio}"

class WallpapersSerializer(serializers.ModelSerializer):
    """壁纸详情序列化器（包含完整信息）"""
    category = serializers.SerializerMethodField()
    tags = serializers.SerializerMethodField()
    aspect_ratio = serializers.SerializerMethodField()
    is_liked = serializers.SerializerMethodField()
    is_collected = serializers.SerializerMethodField()
    uploader = serializers.SerializerMethodField()

    class Meta:
        model = Wallpapers
        fields = [
            'id', 'name', 'url', 'thumb_url', 'width', 'height', 'image_format',
            'source_url', 'description', 'has_watermark', 'category', 'tags',
            'is_live', 'is_hd', 'hot_score', 'like_count', 'collect_count', 'download_count',
            'view_count', 'created_at', 'aspect_ratio', 'is_liked', 'is_collected',
            'uploader',
        ]
        read_only_fields = ['id', 'created_at', 'like_count', 'collect_count', 'download_count']

    def __init__(self, *args, **kwargs):
        super(WallpapersSerializer, self).__init__(*args, **kwargs)
        if not self.context.get('include_detail_info'):
            self.fields.pop('uploader', None)

    def get_uploader(self, obj):
        """获取上传者简要信息"""
        logger.debug(
            f"get_uploader called for wallpaper {obj.id}, include_detail_info: {self.context.get('include_detail_info')}")
        try:
            upload_record = getattr(obj, 'customer_upload', None)
            if not upload_record:
                return None
            customer = getattr(upload_record, 'customer', None)
            if not customer:
                return None
            return {
                "id": customer.id,
                "email": customer.email,
                "nickname": customer.nickname,
                "gender": customer.gender,
                "avatar_url": customer.avatar_url,
                "badge": customer.badge,
                "points": customer.points,
                "level": customer.level,
                "upload_count": customer.upload_count,
                "collection_count": customer.collection_count,
                "last_login": customer.last_login.isoformat() if customer.last_login else None,
                "created_at": customer.created_at.isoformat() if customer.created_at else None,
            }
        except ObjectDoesNotExist:
            return None

    def get_category(self, obj):
        return [
            {
                'id': cat.id,
                'name': cat.name,
            }
            for cat in obj.category.all()
        ]

    def get_tags(self, obj):
        return [
            {
                'id': tag.id,
                'name': tag.name,
            }
            for tag in obj.tags.all()
        ]

    def get_aspect_ratio(self, obj):
        """计算宽高比，格式如 16:9"""
        if not obj.width or not obj.height:
            return None

        def gcd(a, b):
            while b:
                a, b = b, a % b
            return a

        common_divisor = gcd(obj.width, obj.height)
        width_ratio = obj.width // common_divisor
        height_ratio = obj.height // common_divisor
        return f"{width_ratio}:{height_ratio}"

    def get_is_liked(self, obj):
        cid = self.context.get("customer_id")
        if not cid:
            return False
        liked_cache = self.context.get("liked_wallpaper_ids")
        if liked_cache is not None:
            return obj.id in liked_cache
        return WallpaperLike.objects.filter(customer_id=cid, wallpaper_id=obj.pk).exists()

    def get_is_collected(self, obj):
        cid = self.context.get("customer_id")
        if not cid:
            return False
        collected_cache = self.context.get("collected_wallpaper_ids")
        if collected_cache is not None:
            return obj.id in collected_cache
        return WallpaperCollection.objects.filter(user_id=cid, wallpaper_id=obj.pk).exists()


# ======================优化结束end================================
class CollectionItemSerializer(serializers.ModelSerializer):
    wallpaper = WallpapersSerializer(read_only=True)
    uploader = serializers.SerializerMethodField(help_text="上传者信息")

    class Meta:
        model = WallpaperCollection
        fields = ["id", "created_at", "wallpaper", "uploader"]

    def get_uploader(self, obj):
        """获取壁纸的上传者信息"""
        try:
            # 通过 wallpaper_id 查询 CustomerWallpaperUpload
            upload_record = CustomerWallpaperUpload.objects.get(
                wallpaper_id=obj.wallpaper_id
            )
            customer = upload_record.customer
            return {
                'id': customer.id,
                'nickname': customer.nickname,
                'email': customer.email,
                'avatar_url': customer.avatar_url
            }
        except CustomerWallpaperUpload.DoesNotExist:
            # 如果没有上传记录，返回空
            return None

@extend_schema(tags=["壁纸管理"])
@extend_schema_view(
    list=extend_schema(
        summary="获取壁纸列表(Admin也可用)",
        description="默认只显示审核通过或未审核的壁纸（排除审核不通过的）",
        parameters=[
            OpenApiParameter(name="currentPage", type=int, required=False, description="当前页码"),
            OpenApiParameter(name="pageSize", type=int, required=False, description="每页数量"),
            OpenApiParameter(name="name", type=str, required=False, description="壁纸名称搜索"),
            OpenApiParameter(name="tag_id", type=str, required=False,
                             description="标签 ID 查询，支持单个或多个（逗号分隔）"),
            OpenApiParameter(name="tag_name", type=str, required=False,
                             description="标签名称搜索"),
            OpenApiParameter(name="category_id", type=str, required=False,
                             description="分类 ID 筛选，支持单个或多个（逗号分隔）"),
            OpenApiParameter(name="media_live", type=str, required=False, description="静态false或动态true"),
            OpenApiParameter(name="platform", type=str, required=False, description="平台电脑PC或手机PHONE "),
            OpenApiParameter(name="resolution", type=str, required=False,
                             description="分辨率多选，逗号分隔，如 3840x2160,2560x1440,1920x1080"),
            OpenApiParameter(name="aspect_ratio", type=str, required=False,
                             description="宽高比多选，逗号分隔，如 16:9,21:9,9:16"),
            OpenApiParameter(name="order", type=str, required=False,
                             description="排序规则（只能传一个）：latest=最新, views=最多浏览, downloads=最多下载, hot=热度"),
            OpenApiParameter(name="audit_status", type=str, required=False,
                             description="审核状态筛选（仅管理员）：pending/approved/rejected"),

        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "data": {
                        "type": "object",
                        "properties": {
                            "results": {"type": "array", "items": {"$ref": "#/components/schemas/Wallpapers"}},
                            "total": {"type": "integer", "example": 50}
                        }
                    },
                    "message": {"type": "string", "example": "列表获取成功"}
                }
            }
        }
    ),
    retrieve=extend_schema(summary="获取壁纸详情", responses={200: WallpapersSerializer, 404: "壁纸不存在"}),
    create=extend_schema(summary="创建壁纸", request=WallpapersSerializer),
    # update=extend_schema(summary="更新壁纸(Admin或自己上传可删)", request=WallpapersSerializer),
    partial_update=extend_schema(summary="部分更新壁纸", request=WallpapersSerializer),
    destroy=extend_schema(summary="删除壁纸(Admin或自己上传可删)", description="删除指定壁纸记录",
                          responses={204: "删除成功", 404: "壁纸不存在"})
)
class WallpapersViewSet(BaseViewSet):
    """
    壁纸管理 ViewSet
    提供壁纸的增删改查功能
    """
    queryset = Wallpapers.objects.all()
    serializer_class = WallpapersSerializer
    pagination_class = CustomPagination

    def get_permissions(self):
        """根据不同操作返回不同的权限类"""
        if self.action in ['update', 'partial_update', 'destroy']:
            # 写操作：需要是管理员或者是上传者本人
            return [IsOwnerOrAdmin()]
        elif self.action in ['audit_approve', 'audit_reject', 'batch_delete']:
            return [IsAdmin()]
        # 读操作无需权限
        return []
    def get_serializer_class(self):
        """根据动作返回不同的序列化器"""
        if self.action == 'list':
            return WallpapersListSerializer
        return WallpapersSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["customer_id"] = None
        tok = self.request.headers.get("token")
        if tok:
            ok, cid = CustomTokenTool.verify_customer_token(tok)
            if ok:
                ctx["customer_id"] = cid
        return ctx

    def _apply_strategy_sorting(self, queryset, order, platform, page_num, page_size):
        """
        应用推荐策略（仅对普通用户生效）
        order: 'hot' 或 'home'
        返回应用策略后的 queryset
        """
        from django.utils import timezone
        from django.db.models import Value, IntegerField
        from models.models import RecommendStrategy, StrategyWallpaperRelation

        if order not in ['hot', 'home']:
            return queryset

        now = timezone.now()
        # 查找策略：优先平台，其次 all
        matched_strategy = None
        platforms = [platform.lower()] if platform in ['PC', 'PHONE'] else ['all']
        platforms.append('all')  # 确保 all 作为 fallback

        for p in set(platforms):
            strategies = RecommendStrategy.objects.filter(
                platform=p, strategy_type=order, status='active'
            ).order_by('-priority', '-created_at')
            for s in strategies:
                if s.start_time and now < s.start_time:
                    continue
                if s.end_time and now > s.end_time:
                    continue
                matched_strategy = s
                break
            if matched_strategy:
                break

        if not matched_strategy:
            # 无策略，按默认排序
            if order == 'hot':
                return queryset.order_by('-hot_score', '-created_at')
            else:  # home
                return queryset.order_by('-created_at')

        # 获取策略关联的壁纸 ID（按 sort_order, created_at 排序）
        strategy_ids = list(
            StrategyWallpaperRelation.objects.filter(strategy=matched_strategy)
            .order_by('sort_order', '-created_at')
            .values_list('wallpaper_id', flat=True)
        )
        if matched_strategy.content_limit and matched_strategy.content_limit > 0:
            strategy_ids = strategy_ids[:matched_strategy.content_limit]

        strategy_count = len(strategy_ids)
        start_idx = (page_num - 1) * page_size

        # 如果策略数据已展示完，直接返回正常排序
        if start_idx >= strategy_count:
            if order == 'hot':
                return queryset.order_by('-hot_score', '-created_at')
            else:
                return queryset.order_by('-created_at')

        # 构造策略查询集（权重0）
        strategy_qs = Wallpapers.objects.filter(id__in=strategy_ids).annotate(
            _sort_weight=Value(0, output_field=IntegerField())
        )
        # 正常查询集（权重1），排除策略中的壁纸
        normal_qs = queryset.exclude(id__in=strategy_ids).annotate(
            _sort_weight=Value(1, output_field=IntegerField())
        )
        if order == 'hot':
            normal_qs = normal_qs.order_by('-hot_score', '-like_count', '-created_at')
        else:
            normal_qs = normal_qs.order_by('-created_at')

        # 合并并排序（权重0的在前面）
        return strategy_qs.union(normal_qs).order_by('_sort_weight')

    def get_queryset(self):
        """
        基础查询集：只处理审核状态过滤（管理员/普通用户通用）
        """
        from models.models import User
        queryset = super().get_queryset()
        is_admin = False
        if hasattr(self.request, 'user') and isinstance(self.request.user, User):
            if self.request.user.role in ['admin', 'operator', 'super_admin']:
                is_admin = True

        if not is_admin:
            queryset = queryset.exclude(audit_status='rejected')
        else:
            audit_status = self.request.query_params.get("audit_status", "").strip()
            if audit_status and audit_status in ['pending', 'approved', 'rejected']:
                queryset = queryset.filter(audit_status=audit_status)

        return queryset

    def list(self, request, *args, **kwargs):
        """
        壁纸列表接口：管理员与普通用户完全分离
        """
        from models.models import User
        is_admin = hasattr(request, 'user') and isinstance(request.user, User) and \
                   request.user.role in ['admin', 'operator', 'super_admin']
        customer_id = self.get_serializer_context().get("customer_id")

        if is_admin:
            return self._list_for_admin(request, customer_id)
        else:
            return self._list_for_customer(request, customer_id)

    def _list_for_admin(self, request, customer_id):
        """管理员列表逻辑：包含完整字段、不应用策略，按参数直接排序"""
        queryset = self.filter_queryset(self.get_queryset())

        # ---- 筛选条件（与普通用户保持一致）----
        platform = request.query_params.get("platform", "").upper()
        if platform == 'PC':
            queryset = queryset.filter(category__id=1).distinct()
        elif platform == 'PHONE':
            queryset = queryset.filter(category__id=2).distinct()

        tag_name = request.query_params.get("tag_name", "").strip()
        if tag_name:
            tag_ids = list(WallpaperTag.objects.filter(name__icontains=tag_name).values_list('id', flat=True))
            if tag_ids:
                queryset = queryset.filter(tags__id__in=tag_ids).distinct()

        tag_ids_param = request.query_params.get("tag_id")
        if tag_ids_param:
            ids = [int(t) for t in tag_ids_param.split(',') if t.strip().isdigit()]
            if ids:
                queryset = queryset.filter(tags__id__in=ids).distinct()

        # ---- 排序（管理员不使用策略，直接按参数排序）----
        order = request.query_params.get("order", "").lower()
        order_mapping = {
            "hot": "-hot_score",
            "latest": "-created_at",
            "views": "-view_count",
            "downloads": "-download_count",
        }
        if order in order_mapping:
            queryset = queryset.order_by(order_mapping[order], '-created_at')
        else:
            queryset = queryset.order_by('-hot_score', '-created_at')

        # ---- 预加载数据 ----
        queryset = queryset.prefetch_related('tags', 'category').select_related('customer_upload__customer')

        # ---- 分页与序列化 ----
        page = self.paginate_queryset(queryset)
        if page:
            serializer = WallpapersAdminListSerializer(page, many=True, context=self.get_serializer_context())
            return self.get_paginated_response(serializer.data)
        serializer = WallpapersAdminListSerializer(queryset, many=True, context=self.get_serializer_context())
        return ApiResponse(serializer.data)

    def _list_for_customer(self, request, customer_id):
        """普通用户列表逻辑：轻量级字段、支持策略排序、点赞/收藏状态"""
        queryset = self.filter_queryset(self.get_queryset())

        # ---- 筛选条件 ----
        platform = request.query_params.get("platform", "").upper()
        if platform == 'PC':
            queryset = queryset.filter(category__id=1).distinct()
        elif platform == 'PHONE':
            queryset = queryset.filter(category__id=2).distinct()

        tag_name = request.query_params.get("tag_name", "").strip()
        if tag_name:
            tag_ids = list(WallpaperTag.objects.filter(name__icontains=tag_name).values_list('id', flat=True))
            if tag_ids:
                queryset = queryset.filter(tags__id__in=tag_ids).distinct()

        tag_ids_param = request.query_params.get("tag_id")
        if tag_ids_param:
            ids = [int(t) for t in tag_ids_param.split(',') if t.strip().isdigit()]
            if ids:
                queryset = queryset.filter(tags__id__in=ids).distinct()
        media_live = request.query_params.get("media_live", "").lower()
        if media_live == 'true':
            queryset = queryset.filter(category__id=4).distinct()
        elif media_live == 'false':
            queryset = queryset.filter(category__id=3).distinct()
        resolution = request.query_params.get("resolution", "")
        if resolution:
            from django.db.models import Q
            q = Q()
            for res in resolution.split(','):
                if 'x' in res:
                    try:
                        w, h = map(int, res.split('x'))
                        q |= Q(width=w, height=h)
                    except:
                        pass
            if q:
                queryset = queryset.filter(q).distinct()
        # ---- 名称/标签映射筛选 ----
        user_input = request.query_params.get("name", "").strip()
        if user_input:
            matched_tags = []
            for keyword, tag_id in TAG_MAPPING.items():
                if keyword.lower() in user_input.lower():
                    matched_tags.append((len(keyword), keyword, tag_id))
            if matched_tags:
                matched_tags.sort(key=lambda x: x[0], reverse=True)
                tag_id = matched_tags[0][2]
                queryset = queryset.filter(tags__id=tag_id).distinct()
            else:
                queryset = queryset.filter(name__icontains=user_input)
        # ---- 排序（支持策略）----
        order = request.query_params.get("order", "").lower()
        page_num = int(request.query_params.get("currentPage", 1))
        page_size = int(request.query_params.get("pageSize", 10))
        queryset = self._apply_strategy_sorting(queryset, order, platform, page_num, page_size)
        # ---- 预加载轻量级数据 ----
        queryset = queryset.prefetch_related('tags').only(
            'id', 'name', 'url', 'thumb_url', 'width', 'height', 'image_format',
            'has_watermark', 'is_live', 'is_hd', 'hot_score', 'like_count',
            'collect_count', 'download_count', 'view_count', 'created_at', 'audit_status'
        )
        # ---- 用户互动信息 ----
        liked_ids = collected_ids = set()
        if customer_id:
            liked_ids = set(
                WallpaperLike.objects.filter(customer_id=customer_id).values_list('wallpaper_id', flat=True))
            collected_ids = set(
                WallpaperCollection.objects.filter(user_id=customer_id).values_list('wallpaper_id', flat=True))
        # ---- 分页与序列化 ----
        page = self.paginate_queryset(queryset)
        context = self.get_serializer_context()
        context['liked_wallpaper_ids'] = liked_ids
        context['collected_wallpaper_ids'] = collected_ids
        if page:
            serializer = WallpapersListSerializer(page, many=True, context=context)
            return self.get_paginated_response(serializer.data)
        serializer = WallpapersListSerializer(queryset, many=True, context=context)
        return ApiResponse(serializer.data)


    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            wallpaper_id = instance.id
            # 在删除前，先扣减相关用户计数
            _decrement_user_counters_on_wallpaper_delete([wallpaper_id])
            # 执行物理删除
            self.perform_destroy(instance)
            log_operation(
                operator=request.user,
                module="壁纸管理",
                operation_type="delete",
                target_id=wallpaper_id,
                target_name=instance.name,
                description=f"删除壁纸：{instance.name}",
                request=request
            )
            return ApiResponse(message="删除成功")
        except ObjectDoesNotExist:
            return ApiResponse(code=404, message=f"{self.queryset.model.__name__}不存在")
        except Exception as e:
            logger.error(f"删除壁纸失败: {e}", exc_info=True)
            return ApiResponse(code=500, message=f"删除失败: {str(e)}")

    def retrieve(self, request, *args, **kwargs):
        """
        获取壁纸详情，并自动增加浏览量
        """
        instance = self.get_object()
        Wallpapers.objects.filter(pk=instance.pk).update(
            view_count=F("view_count") + 1,
            hot_score=F("hot_score") + 10
        )
        instance.refresh_from_db(fields=["view_count", "hot_score"])
        ctx = self.get_serializer_context()
        ctx['include_detail_info'] = True
        serializer = self.get_serializer(instance, context=ctx)
        return ApiResponse(serializer.data)

    @extend_schema(
        summary="更新壁纸信息",
        description=(
            "管理员或上传者可以更新壁纸信息。"
            "如果 is_change=true 且提供 file，则更换图片并更新所有相关属性；"
            "否则只更新元数据字段（名称、描述、分类、标签等）。"
        ),
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "is_change": {"type": "boolean", "description": "是否更换图片（true=更换图片，false或不传=仅更新元数据）"},
                    "file": {"type": "string", "format": "binary", "description": "新图片文件（is_change=true 时必填）"},
                    "name": {"type": "string", "description": "壁纸名称"},
                    "description": {"type": "string", "description": "壁纸描述"},
                    "source_url": {"type": "string", "description": "来源链接"},
                    "has_watermark": {"type": "boolean", "description": "是否有水印"},
                    "is_hd": {"type": "boolean", "description": "是否高清"},
                    "is_live": {"type": "boolean", "description": "是否动态壁纸"},
                    "category_ids": {
                        "type": "string",
                        "description": "分类ID列表，逗号分隔，如 '1,3'"
                    },
                    "tag_ids": {
                        "type": "string",
                        "description": "标签ID列表，逗号分隔，如 '3677,3680'"
                    },
                    "audit_status": {
                        "type": "string",
                        "enum": ["pending", "approved", "rejected"],
                        "description": "审核状态（仅管理员可修改）"
                    },
                    "audit_remark": {"type": "string", "description": "审核备注（仅管理员）"}
                }
            }
        },
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "data": {"$ref": "#/components/schemas/Wallpapers"},
                    "message": {"type": "string", "example": "更新成功"}
                }
            },
            400: "参数错误",
            403: "无权限",
            404: "壁纸不存在"
        }
    )
    def update(self, request, *args, **kwargs):
        """
        更新壁纸信息
        - 如果 is_change=true 且提供 file：更换图片并更新所有属性
        - 否则：只更新元数据字段
        - 管理员可修改审核状态
        """
        from django.db import transaction

        instance = self.get_object()

        # 判断是否更换图片
        is_change_param = request.data.get("is_change", "false")
        if isinstance(is_change_param, bool):
            is_change = is_change_param
        else:
            is_change = str(is_change_param).lower() == "true"

        # 如果是非管理员，不允许修改审核相关字段
        is_admin = False
        if hasattr(request, 'user') and request.user:
            from models.models import User
            if isinstance(request.user, User) and request.user.role in ['admin', 'operator', 'super_admin']:
                is_admin = True

        if not is_admin:
            request.data.pop('audit_status', None)
            request.data.pop('audit_remark', None)
            request.data.pop('audited_at', None)

        try:
            with transaction.atomic():
                if is_change:
                    # === 更换图片模式 ===
                    uploaded_file = request.FILES.get("file")
                    if not uploaded_file:
                        return ApiResponse(code=400, message="更换图片时必须上传文件 file")

                    # 处理文件上传
                    orig_name = uploaded_file.name or "image.jpg"
                    _, orig_ext = os.path.splitext(orig_name)
                    ext = orig_ext.lower()
                    if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".mp4"):
                        ext = ".jpg"

                    # 生成唯一的文件名
                    timestamp = timezone.now().strftime("%Y%m%d%H%M%S")
                    unique_id = uuid.uuid4().hex[:8]
                    unique_base = f"{timestamp}_{unique_id}"

                    # COS 路径
                    cos_key = f"wallpaper/{unique_base}{ext}"
                    thumb_cos_key = f"wallpaper/{unique_base}_thumb{ext}"
                    _ext_hint = ext.lstrip(".")

                    # 读取文件内容
                    try:
                        file_content = bytes_from_uploaded_image(uploaded_file, quality=100)
                    except Exception as e:
                        logger.error(f"读取上传文件失败: {e}", exc_info=True)
                        return ApiResponse(code=400, message=f"读取文件失败：{e}")

                    # 上传原图到 COS
                    cos_ret = upload_image_to_cos(file_content, cos_key)
                    if not cos_ret:
                        return ApiResponse(code=500, message="上传原图到云存储失败")

                    file_url = cos_ret["url"]

                    # 处理缩略图
                    if ext == ".mp4":
                        thumb_url = ""
                        w, h = 0, 0
                        pil_fmt = "mp4"
                    else:
                        try:
                            uploaded_file.seek(0)
                            thumb_content = bytes_from_uploaded_image(uploaded_file, quality=10)
                            thumb_ret = upload_image_to_cos(thumb_content, thumb_cos_key)
                            if thumb_ret:
                                thumb_url = thumb_ret["url"]
                            else:
                                thumb_url = file_url.rsplit(".", 1)[0] + "_thumb." + _ext_hint

                            w, h, pil_fmt = _image_meta_from_bytes(file_content)
                        except Exception as e:
                            logger.error(f"生成缩略图失败: {e}", exc_info=True)
                            thumb_url = file_url.rsplit(".", 1)[0] + "_thumb." + _ext_hint
                            w, h, pil_fmt = 0, 0, None

                    # 格式化图片格式
                    fmt = (pil_fmt or _ext_hint or "").lower()
                    if fmt == "jpeg":
                        fmt = "jpg"

                    # 自动判断是否高清
                    is_hd_param = request.data.get("is_hd")
                    if is_hd_param is not None:
                        new_is_hd = str(is_hd_param).lower() == "true"
                    else:
                        new_is_hd = w >= 1920 or h >= 1080 if (w and h) else False

                    # 自动判断是否动态壁纸
                    is_live_param = request.data.get("is_live")
                    if is_live_param is not None:
                        new_is_live = str(is_live_param).lower() == "true"
                    else:
                        new_is_live = (ext == ".mp4")

                    # 更新图片相关字段
                    instance.url = file_url[:500]
                    instance.thumb_url = thumb_url[:500] if thumb_url else None
                    instance.width = w or 0
                    instance.height = h or 0
                    instance.image_format = fmt[:20] if fmt else None
                    instance.is_hd = new_is_hd
                    instance.is_live = new_is_live

                # 更新其他字段（无论是否更换图片都可以更新）
                name = request.data.get("name")
                description = request.data.get("description")
                source_url = request.data.get("source_url")
                has_watermark = request.data.get("has_watermark")
                audit_status = request.data.get("audit_status")
                audit_remark = request.data.get("audit_remark")

                if name is not None:
                    instance.name = str(name)[:200]
                if description is not None:
                    instance.description = str(description).strip() or None
                if source_url is not None:
                    instance.source_url = str(source_url)[:500] or None
                if has_watermark is not None and not is_change:
                    # 如果没换图片，可以手动修改 has_watermark
                    instance.has_watermark = str(has_watermark).lower() == "true"
                elif has_watermark is not None and is_change:
                    # 如果换了图片，也可以手动覆盖
                    instance.has_watermark = str(has_watermark).lower() == "true"

                if audit_status is not None:
                    if audit_status not in ['pending', 'approved', 'rejected']:
                        return ApiResponse(code=400, message="审核状态无效")
                    instance.audit_status = audit_status
                    instance.audited_at = timezone.now() if audit_status in ['approved', 'rejected'] else None
                if audit_remark is not None:
                    instance.audit_remark = str(audit_remark).strip() or None

                instance.save()
                # 提取分类和标签ID
                category_ids = request.data.get("category_ids")
                tag_ids = request.data.get("tag_ids")
                # 更新分类（如果提供）
                if category_ids is not None:
                    instance.category.clear()
                    if isinstance(category_ids, list):
                        for cid in category_ids:
                            try:
                                instance.category.add(int(cid))
                            except (ValueError, TypeError):
                                pass
                    elif isinstance(category_ids, str):
                        cat_id_list = []
                        clean_str = category_ids.strip().strip('[]')
                        for cid in clean_str.split(','):
                            cid_clean = cid.strip()
                            if cid_clean:
                                try:
                                    cat_id_list.append(int(cid_clean))
                                except ValueError:
                                    pass
                        for cid in cat_id_list:
                            instance.category.add(cid)
                # 更新标签（如果提供）
                if tag_ids is not None:
                    instance.tags.clear()
                    if isinstance(tag_ids, list):
                        for tid in tag_ids:
                            try:
                                instance.tags.add(int(tid))
                            except (ValueError, TypeError):
                                pass
                    elif isinstance(tag_ids, str):
                        tag_id_list = []
                        clean_str = tag_ids.strip().strip('[]')
                        for tid in clean_str.split(','):
                            tid_clean = tid.strip()
                            if tid_clean:
                                try:
                                    tag_id_list.append(int(tid_clean))
                                except ValueError:
                                    pass
                        for tid in tag_id_list:
                            instance.tags.add(tid)
                # 刷新实例以获取最新数据
                instance.refresh_from_db()

                # 返回更新后的完整数据
                ctx = self.get_serializer_context()
                ctx['include_detail_info'] = True
                response_serializer = WallpapersSerializer(instance, context=ctx)
                action_desc = "更换图片并更新" if is_change else "更新信息"
                log_operation(
                    operator=request.user,
                    module="壁纸管理",
                    operation_type="update",
                    target_id=instance.id,
                    target_name=instance.name,
                    description=f"{action_desc}壁纸：{instance.name}",
                    request=request
                )
                return ApiResponse(
                    data=response_serializer.data,
                    message="更换图片并更新成功" if is_change else "更新成功"
                )
        except Exception as e:
            logger.error(f"更新壁纸失败: {e}", exc_info=True)
            return ApiResponse(code=500, message=f"更新失败：{str(e)}")

    @extend_schema(
        summary="审核通过单张壁纸(Admin)",
        request=None,
        responses={200: {"type": "object", "properties": {"code": {"type": "integer"}, "message": {"type": "string"}}}}
    )
    @action(detail=True, methods=['post'], url_path='audit/approve')
    def audit_approve(self, request, pk=None):
        """
        审核通过壁纸
        仅管理员可用
        """
        from django.utils import timezone
        try:
            wallpaper = Wallpapers.objects.get(id=pk)
        except Wallpapers.DoesNotExist:
            return ApiResponse(code=404, message="壁纸不存在")
        wallpaper.audit_status = 'approved'
        wallpaper.audited_at = timezone.now()
        wallpaper.audit_remark = request.data.get('remark', '')
        wallpaper.save(update_fields=['audit_status', 'audited_at', 'audit_remark'])
        logger.info(f"壁纸 #{pk} 审核通过 by {request.user.username}")
        log_operation(
            operator=request.user,
            module="壁纸管理",
            operation_type="audit",
            target_id=pk,
            target_name=wallpaper.name,
            description=f"审核通过壁纸：{wallpaper.name}",
            request=request
        )
        return ApiResponse(message="审核通过")

    @extend_schema(
        summary="审核拒绝单张(Admin)",
        request=None,
        responses={200: {"type": "object", "properties": {"code": {"type": "integer"}, "message": {"type": "string"}}}}
    )
    @action(detail=True, methods=['post'], url_path='audit/reject')
    def audit_reject(self, request, pk=None):
        """
        审核拒绝壁纸
        仅管理员可用
        """
        from django.utils import timezone
        try:
            wallpaper = Wallpapers.objects.get(id=pk)
        except Wallpapers.DoesNotExist:
            return ApiResponse(code=404, message="壁纸不存在")

            # 获取拒绝原因
        remark = request.data.get('remark', '')
        wallpaper.audit_status = 'rejected'
        wallpaper.audited_at = timezone.now()
        wallpaper.audit_remark = remark
        wallpaper.save(update_fields=['audit_status', 'audited_at', 'audit_remark'])
        logger.info(f"壁纸 #{pk} 审核拒绝 by {request.user.username}, 原因: {remark}")
        log_operation(
            operator=request.user,
            module="壁纸管理",
            operation_type="audit",
            target_id=pk,
            target_name=wallpaper.name,
            description=f"审核拒绝壁纸：{wallpaper.name}，原因：{remark}",
            request=request
        )
        return ApiResponse(message="审核拒绝")

    @extend_schema(
        summary="批量审核通过(Admin)",
        request=OpenApiRequest(
            request={"wallpaper_ids": [1, 2, 3]},
            examples=[
                OpenApiExample(
                    name="审核通过示例",
                    value={"wallpaper_ids": [805471, 805472, 805473]},
                    request_only=True
                )
            ]
        ),
        responses={200: {"type": "object", "properties": {"code": {"type": "integer"}, "data": {"type": "object"},
                                                          "message": {"type": "string"}}}}
    )
    @action(detail=False, methods=['post'], url_path='audit/batch-approve')
    def audit_batch_approve(self, request):
        """
        批量审核通过
        仅管理员可用
        """
        from django.utils import timezone
        wallpaper_ids = request.data.get('wallpaper_ids', [])
        if not wallpaper_ids:
            return ApiResponse(code=400, message="请提供壁纸ID列表")
        count = Wallpapers.objects.filter(id__in=wallpaper_ids).update(
            audit_status='approved',
            audited_at=timezone.now()

        )
        logger.info(f"批量审核通过 {count} 张壁纸 by {request.user.username}")
        log_operation(
            operator=request.user,
            module="壁纸管理",
            operation_type="audit",
            target_id=",".join(map(str, wallpaper_ids)),
            description=f"批量审核通过 {count} 张壁纸",
            request=request,
            extra_data={"wallpaper_ids": wallpaper_ids}
        )
        return ApiResponse(data={'count': count}, message=f"已审核通过 {count} 张壁纸")

    @extend_schema(
        summary="批量审核拒绝(Admin)",
        request=OpenApiRequest(
            request={"wallpaper_ids": [1, 2, 3], "remark": "拒绝原因"},
            examples=[
                OpenApiExample(
                    name="审核拒绝示例",
                    value={"wallpaper_ids": [805471, 805472, 805473], "remark": "图片模糊不清"},
                    request_only=True
                )
            ]
        ),
        responses={200: {"type": "object", "properties": {"code": {"type": "integer"}, "data": {"type": "object"},
                                                          "message": {"type": "string"}}}}
    )
    @action(detail=False, methods=['post'], url_path='audit/batch-reject')
    def audit_batch_reject(self, request):
        """
        批量审核拒绝
        仅管理员可用
        """
        from django.utils import timezone
        wallpaper_ids = request.data.get('wallpaper_ids', [])
        remark = request.data.get('remark', '')

        count = Wallpapers.objects.filter(id__in=wallpaper_ids).update(
            audit_status='rejected',
            audited_at=timezone.now(),
            audit_remark=remark
        )
        logger.info(f"批量审核拒绝 {count} 张壁纸 by {request.user.username}, 原因: {remark}")
        log_operation(
            operator=request.user,
            module="壁纸管理",
            operation_type="audit",
            target_id=",".join(map(str, wallpaper_ids)),
            description=f"批量审核拒绝 {count} 张壁纸，原因：{remark}",
            request=request,
            extra_data={"wallpaper_ids": wallpaper_ids, "remark": remark}
        )
        return ApiResponse(data={'count': count}, message=f"已审核拒绝 {count} 张壁纸")


    @extend_schema(
        summary="猜你喜欢",
        description="根据传入的壁纸 ID，推荐具有相同标签的其他壁纸（按匹配标签数量排序）",
        parameters=[
            OpenApiParameter(name="wallpaper_id", type=int, required=True, description="壁纸 ID"),
            OpenApiParameter(name="limit", type=int, required=False, description="返回数量限制，默认 10"),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "推荐成功"},
                    "data": {
                        "type": "array",
                        "items": {"$ref": "#/components/schemas/Wallpapers"}
                    }
                }
            },
            404: {"description": "壁纸不存在"}
        }
    )
    @action(detail=False, methods=['get'], url_path='guess_you_like')
    def guess_you_like(self, request):
        """
        猜你喜欢：根据传入的壁纸 ID，推荐具有相同标签的其他壁纸
        """
        from django.db.models import Count, Q
        wallpaper_id = request.query_params.get("wallpaper_id")
        limit = int(request.query_params.get("limit", 10))
        if not wallpaper_id:
            return ApiResponse(code=400, message="请提供壁纸 ID")
        try:
            # 获取指定壁纸及其标签
            target_wallpaper = Wallpapers.objects.prefetch_related('tags').get(id=wallpaper_id)
            target_tags = target_wallpaper.tags.all()
            if not target_tags.exists():
                # 如果该壁纸没有标签，返回空列表
                return ApiResponse(data=[], message="该壁纸没有标签，无法推荐")
            target_tag_ids = list(target_tags.values_list('id', flat=True))
            # 查询具有相同标签的其他壁纸（排除自身）
            # 使用 annotate 计算匹配的标签数量
            recommended_wallpapers = Wallpapers.objects.filter(
                tags__id__in=target_tag_ids
            ).exclude(
                id=wallpaper_id
            ).annotate(
                match_count=Count('tags', filter=Q(tags__id__in=target_tag_ids))
            ).order_by(
                '-match_count', '-hot_score', '-created_at'
            ).distinct()[:limit]
            # 序列化返回数据
            serializer = self.get_serializer(recommended_wallpapers, many=True)
            return ApiResponse(data=serializer.data, message=f"为您找到{len(serializer.data)}个相关推荐")
        except Wallpapers.DoesNotExist:
            return ApiResponse(code=404, message="指定的壁纸不存在")
        except Exception as e:
            logger.error(f"猜你喜欢推荐失败：{str(e)}", exc_info=True)
            return ApiResponse(code=500, message=f"推荐失败：{str(e)}")

    @extend_schema(summary="点赞/取消点赞（需客户 Token）")
    @action(
        detail=False,
        methods=["post"],
        url_path="toggle-like",
        permission_classes=[IsCustomerTokenValid],
    )
    def toggle_like(self, request):
        wid = request.data.get("wallpaper_id")
        if not wid:
            return ApiResponse(code=400, message="请提供 wallpaper_id")
        try:
            wid = int(wid)
        except (TypeError, ValueError):
            return ApiResponse(code=400, message="wallpaper_id 无效")
        cid = request.user.id
        try:
            with transaction.atomic():
                wp = Wallpapers.objects.select_for_update().get(pk=wid)
                like, created = WallpaperLike.objects.get_or_create(
                    customer_id=cid, wallpaper=wp
                )
                if created:
                    Wallpapers.objects.filter(pk=wp.pk).update(
                        like_count=F("like_count") + 1,
                        hot_score=F("hot_score") + 50  # 点赞一次加 50 分
                    )
                    liked = True
                    message = "点赞成功"
                    
                    # 发送点赞通知（如果不是给自己点赞）
                    try:
                        upload_record = getattr(wp, 'customer_upload', None)
                        if upload_record and upload_record.customer_id != cid:
                            from App.view.notifications.notification_center import NotificationCenter
                            NotificationCenter.send_like(
                                recipient_id=upload_record.customer_id,
                                sender_id=cid,
                                wallpaper_id=wp.id,
                                wallpaper_name=wp.name[:50]
                            )
                    except Exception:
                        pass
                else:
                    like.delete()
                    message = "取消点赞成功"
                    Wallpapers.objects.filter(pk=wp.pk).update(
                        like_count=Greatest(F("like_count") - 1, 0),
                        hot_score=Greatest(F("hot_score") - 50, 0)  # 取消点赞减 50 分
                    )
                    liked = False
                wp.refresh_from_db(fields=["like_count"])
        except Wallpapers.DoesNotExist:
            return ApiResponse(code=404, message="壁纸不存在")
        return ApiResponse(
            data={"liked": liked, "like_count": wp.like_count},
            message=message,
        )

    @extend_schema(summary="收藏/取消收藏（需客户 Token）")
    @action(
        detail=False,
        methods=["post"],
        url_path="toggle-collect",
        permission_classes=[IsCustomerTokenValid],
    )
    def toggle_collect(self, request):
        wid = request.data.get("wallpaper_id")
        if not wid:
            return ApiResponse(code=400, message="请提供 wallpaper_id")
        try:
            wid = int(wid)
        except (TypeError, ValueError):
            return ApiResponse(code=400, message="wallpaper_id 无效")
        cid = request.user.id
        try:
            with transaction.atomic():
                wp = Wallpapers.objects.select_for_update().get(pk=wid)
                row, created = WallpaperCollection.objects.get_or_create(
                    user_id=cid, wallpaper=wp
                )
                if created:
                    Wallpapers.objects.filter(pk=wp.pk).update(
                        collect_count=F("collect_count") + 1,
                        hot_score=F("hot_score") + 200  # 收藏一次加 200 分
                    )
                    CustomerUser.objects.filter(pk=cid).update(
                        collection_count=F("collection_count") + 1
                    )
                    collected = True
                else:
                    row.delete()
                    Wallpapers.objects.filter(pk=wp.pk).update(
                        collect_count=Greatest(F("collect_count") - 1, 0),
                        hot_score=Greatest(F("hot_score") - 200, 0)  # 取消收藏减 200 分
                    )
                    CustomerUser.objects.filter(pk=cid).update(
                        collection_count=Greatest(F("collection_count") - 1, 0)
                    )
                    collected = False
                wp.refresh_from_db(fields=["collect_count"])
        except Wallpapers.DoesNotExist:
            return ApiResponse(code=404, message="壁纸不存在")
        return ApiResponse(
            data={"collected": collected, "collect_count": wp.collect_count},
            message="操作成功",
        )
        # 2) 在 WallpapersViewSet 里新增批量删除接口
    @extend_schema(
        summary="批量删除壁纸(Admin)",
        description="根据 wallpaper_ids 批量删除壁纸",
        request={
            "application/json": {
                "type": "object",
                "properties": {
                    "wallpaper_ids": {
                        "type": "array",
                        "items": {"type": "integer"},
                        "description": "壁纸ID列表，如 [1,2,3,4]"
                    }
                },
                "required": ["wallpaper_ids"],
                "example": {"wallpaper_ids": [1, 2, 3, 4]}
            }
        },
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer"},
                    "data": {"type": "object"},
                    "message": {"type": "string"}
                }
            }
        }
    )
    @action(detail=False, methods=['post'], url_path='batch-delete')
    @action(detail=False, methods=['post'], url_path='batch-delete')
    def batch_delete(self, request):
        wallpaper_ids = request.data.get('wallpaper_ids', [])
        if not isinstance(wallpaper_ids, list) or not wallpaper_ids:
            return ApiResponse(code=400, message="wallpaper_ids 必须是非空数组")

        valid_ids = []
        for item in wallpaper_ids:
            try:
                valid_ids.append(int(item))
            except (TypeError, ValueError):
                return ApiResponse(code=400, message="wallpaper_ids 只能包含整数ID")

        # 去重，避免重复删除
        valid_ids = list(set(valid_ids))

        queryset = Wallpapers.objects.filter(id__in=valid_ids)
        deleted_count = queryset.count()
        if deleted_count == 0:
            return ApiResponse(code=404, message="未找到可删除的壁纸")

        # 1. 在删除前，预取相关数据用于日志记录
        wallpapers_to_delete = list(queryset.select_related('customer_upload__customer'))
        wallpaper_names = [wp.name for wp in wallpapers_to_delete]

        # 2. 执行物理删除（在事务中保证原子性）
        with transaction.atomic():
            # 先删除关联的点赞和收藏记录（防止外键约束或脏数据）
            WallpaperLike.objects.filter(wallpaper_id__in=valid_ids).delete()
            WallpaperCollection.objects.filter(wallpaper_id__in=valid_ids).delete()

            # 删除壁纸本身
            queryset.delete()

        # 3. 扣减相关用户计数（复用通用函数）
        counter_result = _decrement_user_counters_on_wallpaper_delete(valid_ids)

        log_operation(
            operator=request.user,
            module="壁纸管理",
            operation_type="delete",
            target_id=",".join(map(str, valid_ids)),
            description=f"批量删除 {deleted_count} 张壁纸",
            request=request,
            extra_data={
                "wallpaper_ids": valid_ids,
                "wallpaper_names": wallpaper_names[:10],  # 只记录前10个名称
                "affected_users": counter_result['affected_users']
            }
        )
        return ApiResponse(
            data={
                "deleted_count": deleted_count,
                "wallpaper_ids": valid_ids,
                "affected_users": counter_result['affected_users']
            },
            message=f"批量删除成功，共删除 {deleted_count} 条，影响 {counter_result['affected_users']} 个用户"
        )

    @extend_schema(summary="记录一次下载并返回累计下载量（需客户 Token）")
    @action(
        detail=False,
        methods=["post"],
        url_path="record-download",
        permission_classes=[IsCustomerTokenValid],
    )
    def record_download(self, request):
        wid = request.data.get("wallpaper_id")
        if not wid:
            return ApiResponse(code=400, message="请提供 wallpaper_id")
        try:
            wid = int(wid)
        except (TypeError, ValueError):
            return ApiResponse(code=400, message="wallpaper_id 无效")
        try:
            with transaction.atomic():
                wp = Wallpapers.objects.select_for_update().get(pk=wid)
                Wallpapers.objects.filter(pk=wp.pk).update(
                    download_count=F("download_count") + 1,
                    hot_score=F("hot_score") + 400
                )
                wp.refresh_from_db(fields=["download_count", "hot_score"])
        except Wallpapers.DoesNotExist:
            return ApiResponse(code=404, message="壁纸不存在")
        return ApiResponse(
            data={"download_count": wp.download_count},
            message="记录成功",
        )
    @extend_schema(
        summary="我的收藏列表（仅需客户 Token）",
        parameters=[
            OpenApiParameter(name="currentPage", type=int, required=False, description="当前页码"),
            OpenApiParameter(name="pageSize", type=int, required=False, description="每页数量"),
            OpenApiParameter(name="platform", type=str, required=False, description="平台筛选：PC 或 PHONE"),
            OpenApiParameter(name="customer_id", type=str, required=False, description="他人id"),
        ],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="my-collections",
        permission_classes=[IsCustomerTokenValid],
    )
    def my_collections(self, request):
        token = request.headers.get("token")
        is_valid, customer_id = CustomTokenTool.verify_customer_token(token)
        target_customer_id = request.query_params.get('customer_id')
        role = False
        if target_customer_id and request.user and hasattr(request.user, 'role'):
            role = True
            # 检查是否为管理员（假设 role 字段存在且为 admin 相关角色）
            try:
                customer_id = int(target_customer_id)
            except Exception as e:
                return ApiResponse(code=400, message="无效的 customer_id")

        user_allowed = is_valid and customer_id
        admin_allowed = role and customer_id
        if not (user_allowed or admin_allowed):
            return ApiResponse(code=401, message="客户 Token 无效或已过期")

        qs = (
            WallpaperCollection.objects
            .filter(user_id=customer_id)
            .select_related("wallpaper")
            .prefetch_related("wallpaper__tags", "wallpaper__category")
            .order_by("-created_at")
        )

        platform = request.query_params.get("platform", "").upper()
        if platform == 'PC':
            qs = qs.filter(wallpaper__category__id=1)
        elif platform == 'PHONE':
            qs = qs.filter(wallpaper__category__id=2)

        page = self.paginate_queryset(qs)
        if page is not None:
            data = CollectionItemSerializer(
                page,
                many=True,
                context=self.get_serializer_context(),
            ).data
            return self.get_paginated_response(data)
        data = CollectionItemSerializer(
            qs,
            many=True,
            context=self.get_serializer_context(),
        ).data
        return ApiResponse(data=data, message="获取收藏列表成功")


    @extend_schema(
        summary="管理员上传壁纸",
        description=(
            "管理员上传壁纸到 COS，自动处理原图和缩略图。"
            "支持设置分类、标签、审核状态等完整字段。"
        ),
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {"type": "string", "format": "binary", "description": "壁纸文件（必填）"},
                    "name": {"type": "string", "description": "壁纸名称（必填）"},
                    "description": {"type": "string", "description": "壁纸描述（可选）"},
                    "source_url": {"type": "string", "description": "来源链接（可选）"},
                    "has_watermark": {"type": "boolean", "description": "是否有水印（可选，默认 false）"},
                    "is_hd": {"type": "boolean", "description": "是否高清（可选，默认自动判断）"},
                    "is_live": {"type": "boolean", "description": "是否动态壁纸（可选，默认根据文件扩展名判断）"},
                    "category_ids": {
                        "type": "string",
                        "description": "分类 ID 列表，逗号分隔，如 '1,3' 表示电脑+静态（可选）"
                    },
                    "tag_ids": {
                        "type": "string",
                        "description": "标签 ID 列表，逗号分隔，如 '3677,3680'（可选）"
                    },
                    "audit_status": {
                        "type": "string",
                        "enum": ["pending", "approved", "rejected"],
                        "description": "审核状态（可选，默认 pending）"
                    },
                    "audit_remark": {"type": "string", "description": "审核备注（可选）"}
                },
                "required": ["file", "name"]
            }
        },
        responses={
            201: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 201},
                    "data": {"$ref": "#/components/schemas/Wallpapers"},
                    "message": {"type": "string", "example": "上传成功"}
                }
            },
            400: "参数错误",
            500: "服务器错误"
        }
    )
    @action(
        detail=False,
        methods=["post"],
        url_path="upload-admin",
        permission_classes=[IsAdmin],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_admin(self, request):
        """
        管理员上传壁纸
        - 自动处理原图和缩略图
        - 支持设置分类、标签、审核状态
        - 需要管理员权限
        """
        from django.db import transaction

        # 验证文件
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return ApiResponse(code=400, message="请上传文件 file")
        # 验证必填字段
        name = (request.data.get("name") or "").strip()
        if not name:
            return ApiResponse(code=400, message="请提供壁纸名称 name")
        # 解析可选字段
        description = (request.data.get("description") or "").strip() or None
        source_url = (request.data.get("source_url") or "").strip() or None
        has_watermark = request.data.get("has_watermark", "false").lower() == "true"
        # 解析 is_hd 和 is_live（可选，不传则自动判断）
        is_hd_param = request.data.get("is_hd")
        is_live_param = request.data.get("is_live")
        audit_status = request.data.get("audit_status", "pending").strip()
        audit_remark = (request.data.get("audit_remark") or "").strip() or None
        # 验证审核状态
        if audit_status and audit_status not in ['pending', 'approved', 'rejected']:
            return ApiResponse(code=400, message="审核状态无效，可选值：pending/approved/rejected")
        # 解析分类和标签
        category_ids = _parse_tag_ids(request.data.get("category_ids"))
        tag_ids = _parse_tag_ids(request.data.get("tag_ids"))
        # 处理文件扩展名
        orig_name = uploaded_file.name or "image.jpg"
        _, orig_ext = os.path.splitext(orig_name)
        ext = orig_ext.lower()
        if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".mp4"):
            ext = ".jpg"
        # 生成唯一的文件名
        timestamp = timezone.now().strftime("%Y%m%d%H%M%S")
        unique_id = uuid.uuid4().hex[:8]
        unique_base = f"{timestamp}_{unique_id}"
        # COS 路径
        cos_key = f"wallpaper/{unique_base}{ext}"
        thumb_cos_key = f"wallpaper/{unique_base}_thumb{ext}"
        _ext_hint = ext.lstrip(".")
        # 读取文件内容
        try:
            file_content = bytes_from_uploaded_image(uploaded_file, quality=100)
        except Exception as e:
            logger.error(f"读取上传文件失败: {e}", exc_info=True)
            return ApiResponse(code=400, message=f"读取文件失败：{e}")
        # 上传原图到 COS
        cos_ret = upload_image_to_cos(file_content, cos_key)
        if not cos_ret:
            return ApiResponse(code=500, message="上传原图到云存储失败，请检查 COS 配置")
        file_url = cos_ret["url"]
        # 处理缩略图
        if ext == ".mp4":
            # 视频文件不生成缩略图
            thumb_url = ""
            w, h = 0, 0
            pil_fmt = "mp4"
        else:
            # 图片文件生成缩略图
            try:
                uploaded_file.seek(0)
                thumb_content = bytes_from_uploaded_image(uploaded_file, quality=10)
                thumb_ret = upload_image_to_cos(thumb_content, thumb_cos_key)
                if thumb_ret:
                    thumb_url = thumb_ret["url"]
                else:
                    thumb_url = file_url.rsplit(".", 1)[0] + "_thumb." + _ext_hint

                # 获取图片尺寸
                w, h, pil_fmt = _image_meta_from_bytes(file_content)
            except Exception as e:
                logger.error(f"生成缩略图失败: {e}", exc_info=True)
                thumb_url = file_url.rsplit(".", 1)[0] + "_thumb." + _ext_hint
                w, h, pil_fmt = 0, 0, None

        # 格式化图片格式
        fmt = (pil_fmt or _ext_hint or "").lower()
        if fmt == "jpeg":
            fmt = "jpg"

        # 自动判断是否高清（如果未手动指定）
        if is_hd_param is not None:
            is_hd = is_hd_param.lower() == "true"
        else:
            is_hd = w >= 1920 or h >= 1080 if (w and h) else False

        # 自动判断是否动态壁纸（如果未手动指定）
        if is_live_param is not None:
            is_live = is_live_param.lower() == "true"
        else:
            is_live = (ext == ".mp4")
        try:
            with transaction.atomic():
                # 创建壁纸记录
                wp = Wallpapers.objects.create(
                    name=name[:200],
                    url=file_url[:500],
                    thumb_url=thumb_url[:500] if thumb_url else None,
                    width=w or 0,
                    height=h or 0,
                    image_format=(fmt[:20] if fmt else None),
                    source_url=source_url[:500] if source_url else None,
                    description=description,
                    has_watermark=has_watermark,
                    is_hd=is_hd,
                    is_live=is_live,
                    audit_status=audit_status if audit_status else 'pending',
                    audit_remark=audit_remark,
                    audited_at=timezone.now() if audit_status in ['approved', 'rejected'] else None
                )

                # 添加分类（如果提供了）
                if category_ids:
                    wp.category.set(category_ids)

                # 添加标签（如果提供了）
                if tag_ids:
                    tag_objs = []
                    for tid in tag_ids:
                        t = WallpaperTag.objects.filter(pk=tid).first()
                        if t:
                            tag_objs.append(t)
                    wp.tags.set(tag_objs)

            # 返回完整的壁纸信息
            ctx = self.get_serializer_context()
            ctx['include_detail_info'] = True
            serializer = WallpapersSerializer(wp, context=ctx)
            log_operation(
                operator=request.user,
                module="壁纸管理",
                operation_type="create",
                target_id=wp.id,
                target_name=wp.name,
                description=f"管理员上传壁纸：{wp.name}",
                request=request
            )
            return ApiResponse(
                data=serializer.data,
                message="上传成功",
                code=201
            )
        except Exception as e:
            logger.error(f"保存壁纸记录失败: {e}", exc_info=True)
            return ApiResponse(code=500, message=f"保存壁纸失败：{e}")

    @extend_schema(
        summary="我的上传列表（仅需客户 Token）/管理员权限+他人id",
        description="查看当前用户上传的所有壁纸记录，支持分页和平台筛选",
        parameters=[
            OpenApiParameter(name="currentPage", type=int, required=False, description="当前页码"),
            OpenApiParameter(name="pageSize", type=int, required=False, description="每页数量"),
            OpenApiParameter(name="platform", type=str, required=False, description="平台筛选：PC 或 PHONE"),
            OpenApiParameter(name="customer_id", type=str, required=False, description="其他用户id"),
        ],
    )
    @action(
        detail=False,
        methods=["get"],
        url_path="my-uploads",
        permission_classes=[IsCustomerTokenValid],
    )
    def my_uploads(self, request):
        token = request.headers.get("token")
        is_valid, customer_id = CustomTokenTool.verify_customer_token(token)

        target_customer_id = request.query_params.get('customer_id')
        role = False

        if target_customer_id and request.user and hasattr(request.user, 'role'):
            role = True
            # 检查是否为管理员（假设 role 字段存在且为 admin 相关角色）
            try:
                customer_id = int(target_customer_id)
            except Exception as e:
                return ApiResponse(code=400, message="无效的 customer_id")
        user_allowed = is_valid and customer_id
        admin_allowed = role and customer_id
        if not (user_allowed or admin_allowed):
            return ApiResponse(code=401, message="客户 Token 无效或已过期")
        qs = (
            CustomerWallpaperUpload.objects
            .filter(customer_id=customer_id)
            .select_related("wallpaper")
            .prefetch_related("wallpaper__tags", "wallpaper__category")
            .order_by("-created_at")
        )

        platform = request.query_params.get("platform", "").upper()
        if platform == 'PC':
            qs = qs.filter(wallpaper__category__id=1)
        elif platform == 'PHONE':
            qs = qs.filter(wallpaper__category__id=2)

        page = self.paginate_queryset(qs)
        if page is not None:
            # 序列化壁纸数据
            wallpapers = [item.wallpaper for item in page]
            serializer = WallpapersSerializer(
                wallpapers,
                many=True,
                context=self.get_serializer_context(),
            )
            return self.get_paginated_response(serializer.data)

        wallpapers = [item.wallpaper for item in qs]
        data = WallpapersSerializer(
            wallpapers,
            many=True,
            context=self.get_serializer_context(),
        ).data
        return ApiResponse(data=data, message="获取上传列表成功")

    @extend_schema(
        summary="精选壁纸（按平台推荐）",
        description="根据平台（PC/手机）返回精选壁纸，支持自定义数量（5-10张）",
        parameters=[
            OpenApiParameter(name="platform", type=str, required=True, description="平台：PC 或 PHONE"),
            OpenApiParameter(name="limit", type=int, required=False, description="返回数量，默认 6，范围 5-10"),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "精选壁纸获取成功"},
                    "data": {
                        "type": "array",
                        "items": {"$ref": "#/components/schemas/Wallpapers"}
                    }
                }
            },
            400: {"description": "参数错误"}
        }
    )
    @action(detail=False, methods=['get'], url_path='featured')
    def featured(self, request):
        """
        精选壁纸：根据平台返回高质量壁纸，优先应用推荐策略
        策略匹配顺序：platform -> all -> 默认平台分类
        """
        platform = request.query_params.get("platform", "").upper()
        if platform not in ['PC', 'PHONE', 'ALL']:
            return ApiResponse(code=400, message="平台参数错误，请输入 PC、PHONE 或 ALL")
        try:
            limit = int(request.query_params.get("limit", 6))
        except (TypeError, ValueError):
            limit = 6
        from django.utils import timezone
        now = timezone.now()
        matched_strategy = None
        if platform in ['PC', 'PHONE']:
            platform_strategies = RecommendStrategy.objects.filter(
                platform=platform.lower(),
                strategy_type="banner",
                status="active",
            ).order_by("-priority", "-created_at")
            for item in platform_strategies:
                if item.start_time and now < item.start_time:
                    continue
                if item.end_time and now > item.end_time:
                    continue
                matched_strategy = item
                break
            if not matched_strategy:
                all_strategies = RecommendStrategy.objects.filter(
                    platform="all",
                    strategy_type="banner",
                    status="active",
                ).order_by("-priority", "-created_at")

                for item in all_strategies:
                    if item.start_time and now < item.start_time:
                        continue
                    if item.end_time and now > item.end_time:
                        continue
                    matched_strategy = item
                    break
        else:
            all_strategies = RecommendStrategy.objects.filter(
                platform="all",
                strategy_type="banner",
                status="active",
            ).order_by("-priority", "-created_at")

            for item in all_strategies:
                if item.start_time and now < item.start_time:
                    continue
                if item.end_time and now > item.end_time:
                    continue
                matched_strategy = item
                break
        if matched_strategy:
            # 通过关联表查询壁纸，按排序权重排序
            relations = StrategyWallpaperRelation.objects.filter(
                strategy=matched_strategy
            ).select_related('wallpaper').order_by('sort_order', '-created_at')

            # 如果有限制数量
            if matched_strategy.content_limit and matched_strategy.content_limit > 0:
                relations = relations[:matched_strategy.content_limit]

            # 提取壁纸对象
            ordered_wallpapers = [rel.wallpaper for rel in relations if rel.wallpaper]

            if ordered_wallpapers:
                serializer = self.get_serializer(ordered_wallpapers, many=True)
                return ApiResponse(
                    data=serializer.data,
                    message="精选壁纸获取成功（来自推荐策略）"
                )
        if platform == 'PC':
            queryset = Wallpapers.objects.filter(
                category__id=1,
                is_hd=True
            ).distinct().order_by('-hot_score', '-like_count', '-created_at')[:limit]
        elif platform == 'PHONE':
            queryset = Wallpapers.objects.filter(
                category__id=2,
                is_hd=True
            ).distinct().order_by('-hot_score', '-like_count', '-created_at')[:limit]
        else:
            queryset = Wallpapers.objects.filter(
                is_hd=True
            ).distinct().order_by('-hot_score', '-like_count', '-created_at')[:limit]

        serializer = self.get_serializer(queryset, many=True)
        return ApiResponse(data=serializer.data, message="精选壁纸获取成功")

    @extend_schema(
        summary="上传个人壁纸到 COS（person_wallpaper/，质量 100）",
        description=(
                "需客户 Token（CToken）。multipart：file、title（文件名主体）；"
                "可选 description；tag_ids（逗号或 JSON 数组）；tag_names（新标签，逗号或 JSON 数组）。"
        ),
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {"type": "string", "format": "binary"},
                    "title": {"type": "string", "description": "展示名/文件名主体（不含扩展名）"},
                    "platform": {"type": "string", "description": "平台：PC 或 PHONE"},
                    "description": {"type": "string"},
                    "tag_ids": {"type": "string", "description": "已有标签 id，如 1,2 或 [1,2]"},
                    "tag_names": {"type": "string", "description": "新标签名称，多个用逗号或 JSON 数组"},
                },
                "required": ["file", "title"],
            }
        },
    )
    @action(
        detail=False,
        methods=["post"],
        url_path="upload-person",
        permission_classes=[IsCustomerTokenValid],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_person(self, request):
        token = request.headers.get("token")
        is_valid, customer_id = CustomTokenTool.verify_customer_token(token)
        if not is_valid or not customer_id:
            return ApiResponse(code=401, message="客户 Token 无效或已过期")

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return ApiResponse(code=400, message="请上传文件 file")

        title = (request.data.get("title") or "").strip()
        if not title:
            return ApiResponse(code=400, message="请提供 title")

        platform = (request.data.get("platform") or "").upper()
        if platform not in ['PC', 'PHONE', 'ALL']:
            return ApiResponse(code=400, message="平台参数错误，请输入 PC 或 PHONE")

        description = (request.data.get("description") or "").strip() or None
        tag_ids = _parse_tag_ids(request.data.get("tag_ids"))
        tag_names = _parse_tag_names(request.data.get("tag_names"))

        orig_name = uploaded_file.name or "image.jpg"
        _, orig_ext = os.path.splitext(orig_name)
        orig_ext = orig_ext.lower()

        token_suffix = token[-8:] if token and len(token) >= 8 else (token or "00000000")[-8:].ljust(8, '0')
        name_part, ext = os.path.splitext(orig_name)
        ext = ext.lower()
        if ext not in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".mp4"):
            ext = ".jpg"

        safe_title = (title or "").strip()
        for ch in '<>:"|?*\\/\x00':
            safe_title = safe_title.replace(ch, "")
        safe_title = safe_title.strip(" .")[:180]
        if not safe_title:
            safe_title = uuid.uuid4().hex[:12]

        unique_base = f"{token_suffix}_{name_part}"
        cos_key = f"person_wallpaper/{unique_base}{ext}"
        thumb_cos_key = f"person_wallpaper/{unique_base}_thumb{ext}"
        _ext_hint = ext.lstrip(".")

        try:
            file_content = bytes_from_uploaded_image(uploaded_file, quality=100)
        except Exception as e:
            logger.error(f"读取上传文件失败: {e}", exc_info=True)
            return ApiResponse(code=400, message=f"读取文件失败：{e}")

        cos_ret = upload_image_to_cos(file_content, cos_key)
        if not cos_ret:
            return ApiResponse(code=500, message="上传到云存储失败，请检查 COS 配置")

        file_url = cos_ret["url"]

        if ext == ".mp4":
            thumb_url = ""
            w, h = 0, 0
            pil_fmt = "mp4"
        else:
            uploaded_file.seek(0)
            thumb_content = bytes_from_uploaded_image(uploaded_file, quality=10)
            thumb_ret = upload_image_to_cos(thumb_content, thumb_cos_key)
            if thumb_ret:
                thumb_url = thumb_ret["url"]
            else:
                thumb_url = file_url.rsplit(".", 1)[0] + "_thumb." + _ext_hint
            w, h, pil_fmt = _image_meta_from_bytes(file_content)

        fmt = (pil_fmt or _ext_hint or "").lower()
        if fmt == "jpeg":
            fmt = "jpg"
        is_hd = w >= 1920 or h >= 1080 if (w and h) else False

        is_live = (ext == ".mp4")
        # 确定平台分类 (1:电脑, 2:手机)
        platform_cat_id = 1 if platform == 'PC' else 2
        # 确定类型分类 (3:静态, 4:动态), 默认 3
        type_cat_id = 4 if is_live else 3

        try:
            with transaction.atomic():
                wp = Wallpapers.objects.create(
                    name=title[:200],
                    url=file_url[:500],
                    thumb_url=thumb_url[:500] if thumb_url else None,
                    width=w or 0,
                    height=h or 0,
                    image_format=(fmt[:20] if fmt else None),
                    description=description,
                    has_watermark=False,
                    is_hd=is_hd,
                    is_live=is_live,
                )
                # 同时添加平台分类和类型分类
                wp.category.add(platform_cat_id, type_cat_id)

                CustomerWallpaperUpload.objects.create(
                    wallpaper=wp,
                    customer_id=customer_id,
                    cos_key=cos_key[:500] if cos_key else None,
                )
                CustomerUser.objects.filter(pk=customer_id).update(
                    upload_count=F("upload_count") + 1
                )
                tag_objs = []
                for tid in tag_ids:
                    t = WallpaperTag.objects.filter(pk=tid).first()
                    if t:
                        tag_objs.append(t)
                for nm in tag_names:
                    nm_clean = nm[:50].strip()
                    if not nm_clean:
                        continue
                    t, _ = WallpaperTag.objects.get_or_create(name=nm_clean)
                    tag_objs.append(t)
                dedup = list({t.id: t for t in tag_objs}.values())
                wp.tags.set(dedup)
        except Exception as e:
            logger.error(f"保存壁纸记录失败: {e}", exc_info=True)
            return ApiResponse(code=500, message=f"保存壁纸失败：{e}")
        data = WallpapersSerializer(
            wp, context=self.get_serializer_context()
        ).data
        return ApiResponse(
            data={
                **data,
                "cos_key": cos_key,
                "url": file_url,
            },
            message="上传成功",
        )

    @extend_schema(
        summary="标签联想/建议（可选关键词）",
        parameters=[
            OpenApiParameter(name="q", type=str, required=False, description="关键词"),
            OpenApiParameter(name="limit", type=int, required=False, description="返回条数，默认 20，最大 50"),
        ],
    )
    @action(detail=False, methods=["get"], url_path="suggest-tags")
    def suggest_tags(self, request):
        q = (request.query_params.get("q") or "").strip()
        try:
            limit = int(request.query_params.get("limit", 20))
        except ValueError:
            limit = 20
        limit = max(1, min(50, limit))
        qs = WallpaperTag.objects.all().order_by("-created_at")
        if q:
            qs = qs.filter(name__icontains=q)
        qs = qs[:limit]
        data = [{"id": t.id, "name": t.name} for t in qs]
        return ApiResponse(data=data, message="ok")

    @extend_schema(
        summary="获取所有壁纸分类",
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "分类获取成功"},
                    "data": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "integer", "example": 1},
                                "name": {"type": "string", "example": "静态"},
                                "desc": {"type": "string", "example": "静态壁纸分类"},
                                "sort": {"type": "integer", "example": 1},
                                "created_at": {"type": "string", "format": "date-time"}
                            }
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['get'], url_path='categories')
    def categories(self, request):
        """
        获取所有壁纸分类列表
        """
        categories = WallpaperCategory.objects.all().order_by('sort', '-created_at')
        data = [
            {
                'id': cat.id,
                'name': cat.name,
                'desc': cat.desc or '',
                'sort': cat.sort,
                'created_at': cat.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            for cat in categories
        ]
        return ApiResponse(data=data, message="分类获取成功")

    @extend_schema(
        summary="批量导入壁纸数据",
        description="通过上传 Excel 文件批量导入壁纸数据",
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {"type": "string", "format": "binary", "description": "Excel 文件 (.xlsx)"}
                },
                "required": ["file"]
            }
        },
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "导入成功"},
                    "data": {
                        "type": "object",
                        "properties": {
                            "success_count": {"type": "integer", "example": 100},
                            "fail_count": {"type": "integer", "example": 0},
                            "errors": {"type": "array", "items": {"type": "string"}}
                        }
                    }
                }
            },
            400: {"description": "文件格式错误或数据异常"},
            500: {"description": "服务器内部错误"}
        }
    )
    @action(detail=False, methods=['post'], url_path='batch-import')
    def batch_import(self, request):
        """
        批量导入壁纸数据
        Excel 格式要求：
        - 第一行：表头（name, url）
        - 从第二行开始：数据行
        """
        from openpyxl import load_workbook

        # 1. 获取上传的文件
        file = request.FILES.get('file')
        if not file:
            return ApiResponse(code=400, message=_("请上传 Excel 文件"))

        # 2. 验证文件格式
        if not file.name.endswith('.xlsx'):
            return ApiResponse(code=400, message=_("仅支持.xlsx 格式的 Excel 文件"))

        try:
            # 3. 读取 Excel 文件
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            # 4. 获取表头
            headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]

            # 5. 验证必要字段
            required_columns = ['name', 'url']
            if not all(col in headers for col in required_columns):
                return ApiResponse(
                    code=400,
                    message=_("Excel 文件缺少必要字段，需要包含：%(cols)s") % {"cols": ", ".join(required_columns)}
                )

            # 6. 定义数据处理函数
            def process_cell_value(value):
                """处理单元格值，转换类型和处理空值"""
                if value is None or str(value).strip().lower() in ['nan', 'n/a', '', 'none']:
                    return None
                return str(value).strip()

            # 7. 批量导入数据
            success_count = 0
            fail_count = 0
            error_messages = []
            created_list = []
            # 从第 2 行开始读取数据
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 将行数据转换为字典
                    row_dict = dict(zip(headers, row))
                    # 处理每个字段的值
                    name = process_cell_value(row_dict.get('name'))
                    url = process_cell_value(row_dict.get('url'))
                    # 验证必要字段
                    if not name:
                        raise ValueError("壁纸名称不能为空")
                    if not url:
                        raise ValueError("壁纸链接不能为空")
                    # 简单的 URL 格式验证
                    if not url.startswith(('http://', 'https://')):
                        raise ValueError("壁纸链接必须是有效的 HTTP/HTTPS URL")
                    # 创建记录
                    wallpaper = Wallpapers.objects.create(
                        name=name,
                        url=url
                    )
                    created_list.append(f"{name}")
                    success_count += 1
                except Exception as e:
                    fail_count += 1
                    error_msg = f"第{row_idx}行导入失败：{str(e)}"
                    error_messages.append(error_msg)
                    logger.error(error_msg)

            # 8. 返回结果
            result_data = {
                "success_count": success_count,
                "fail_count": fail_count,
                "created_count": len(created_list),
                "created_list": created_list[:10],  # 只显示前 10 个
            }
            if error_messages:
                result_data["errors"] = error_messages[:20]  # 只显示前 20 个错误

            message = _("导入完成！成功：%(success)d条，失败：%(fail)d条，新增：%(created)d条") % {
                "success": success_count,
                "fail": fail_count,
                "created": len(created_list)
            }
            return ApiResponse(data=result_data, message=message)
        except Exception as e:
            logger.error(f"批量导入失败：{str(e)}", exc_info=True)
            return ApiResponse(code=500, message=_("批量导入失败：%(error)s") % {"error": str(e)})
