"""
@Project ：CrushCheck
@File    ：view.py
@Author  ：LiangHB
@Date    ：2026/3/3 15:00
@description :
"""
import os

from django.db.models import Q
from openpyxl.reader.excel import load_workbook
from rest_framework import serializers
from rest_framework.decorators import action

from models.models import WeChatPayOrder, WeChatUser, Area
from tool.middleware import logger
from tool.token_tools import _redis
from tool.base_views import BaseViewSet
from tool.utils import CustomPagination, ApiResponse
from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter
from django.utils.translation import gettext as _


class AreaSerializer(serializers.ModelSerializer):
    """省市区序列化器"""
    children_count = serializers.SerializerMethodField(help_text="子区域数量")

    class Meta:
        model = Area
        fields = ['id', 'name', 'code', 'parent', 'children_count']
        read_only_fields = ['id']

    def get_children_count(self, obj):
        """获取子区域数量"""
        return obj.children.count() if hasattr(obj, 'children') else 0


class AreaTreeSerializer(serializers.Serializer):
    """树形结构序列化器"""
    value = serializers.CharField(source='code', help_text="区域代码")
    label = serializers.CharField(source='name', help_text="区域名称")
    children = serializers.ListField(child=serializers.DictField(), help_text="子区域列表")


@extend_schema(tags=["省市区管理"])
@extend_schema_view(
    list=extend_schema(
        summary="获取省市区列表",
        description="支持按父级代码筛选，获取指定级别的区域",
        parameters=[
            OpenApiParameter(name="parent_code", type=str, location=OpenApiParameter.QUERY,
                             description="父级区域代码（如：110000 获取北京市的下级区域）"),
            OpenApiParameter(name="level", type=str, location=OpenApiParameter.QUERY,
                             description="区域级别（province=省级，city=市级，district=区县级）",
                             enum=["province", "city", "district"]),
            OpenApiParameter(name="keyword", type=str, location=OpenApiParameter.QUERY,
                             description="搜索关键词"),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "data": {
                        "type": "array",
                        "items": {"$ref": "#/components/schemas/Area"}
                    }
                }
            }
        }
    ),
    retrieve=extend_schema(summary="获取省市区详情", responses={200: AreaSerializer}),
)
class AreaViewSet(BaseViewSet):
    """省市区ViewSet"""
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    pagination_class = None  # 默认不分页

    def get_queryset(self):
        """动态过滤查询"""
        queryset = super().get_queryset()

        # 筛选参数处理
        parent_code = self.request.query_params.get("parent_code")
        keyword = self.request.query_params.get("keyword")

        # 根据父级代码筛选
        if parent_code:
            try:
                parent = Area.objects.get(code=parent_code)
                queryset = queryset.filter(parent=parent)
            except Area.DoesNotExist:
                queryset = queryset.none()
        else:
            # 如果没有parent_code，默认返回省级（第一级）
            queryset = queryset.filter(parent__isnull=True)

        # 关键词搜索
        if keyword:
            queryset = queryset.filter(
                Q(name__icontains=keyword) | Q(code__icontains=keyword)
            )

        return queryset

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            # 使用自定义分页响应
            return self.get_paginated_response(serializer.data)
        # 如果没有分页，返回普通响应
        serializer = self.get_serializer(queryset, many=True)
        return ApiResponse(serializer.data)

    @extend_schema(
        summary="获取省份列表",
        description="获取所有省级行政区（直辖市、省、自治区等）",
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "查询成功"},
                    "data": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "integer", "example": 1},
                                "name": {"type": "string", "example": "北京市"},
                                "code": {"type": "string", "example": "110000"}
                            }
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['get'], url_path='provinces')
    def get_provinces(self, request):
        """获取所有省份"""
        provinces = self.get_queryset().filter(parent__isnull=True)
        serializer = self.get_serializer(provinces, many=True)
        return ApiResponse(data=serializer.data, message=_("查询成功"))

    @extend_schema(
        summary="获取城市列表",
        description="根据省份代码获取所有城市",
        parameters=[
            OpenApiParameter(name="province_code", type=str, required=True,
                             location=OpenApiParameter.QUERY,
                             description="省份代码，如：110000（北京市）、440000（广东省）"),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "查询成功"},
                    "data": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "integer", "example": 2},
                                "name": {"type": "string", "example": "石家庄市"},
                                "code": {"type": "string", "example": "130100"}
                            }
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['get'], url_path='cities')
    def get_cities(self, request):
        """根据省份代码获取城市"""
        province_code = request.query_params.get('province_code')

        if not province_code:
            return ApiResponse(code=400, message=_("缺少省份代码参数"))

        try:
            province = Area.objects.get(code=province_code)
            cities = Area.objects.filter(parent=province)
            serializer = self.get_serializer(cities, many=True)
            return ApiResponse(data=serializer.data, message=_("查询成功"))
        except Area.DoesNotExist:
            return ApiResponse(code=404, message=_("省份不存在"))

    @extend_schema(
        summary="获取区县列表",
        description="根据城市代码获取所有区县",
        parameters=[
            OpenApiParameter(name="city_code", type=str, required=True,
                             location=OpenApiParameter.QUERY,
                             description="城市代码，如：130100（石家庄市）"),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "查询成功"},
                    "data": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "integer", "example": 3},
                                "name": {"type": "string", "example": "长安区"},
                                "code": {"type": "string", "example": "130102"}
                            }
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['get'], url_path='districts')
    def get_districts(self, request):
        """根据城市代码获取区县"""
        city_code = request.query_params.get('city_code')

        if not city_code:
            return ApiResponse(code=400, message=_("缺少城市代码参数"))

        try:
            city = Area.objects.get(code=city_code)
            districts = Area.objects.filter(parent=city)
            serializer = self.get_serializer(districts, many=True)
            return ApiResponse(data=serializer.data, message=_("查询成功"))
        except Area.DoesNotExist:
            return ApiResponse(code=404, message=_("城市不存在"))

    @extend_schema(
        summary="获取省市区树形结构",
        description="获取完整的三级联动树形结构数据",
        parameters=[
            OpenApiParameter(name="depth", type=int, location=OpenApiParameter.QUERY,
                             description="深度：1=只返回省份，2=返回省市，3=返回省市区（默认）",
                             default=3),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "data": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "value": {"type": "string", "example": "110000"},
                                "label": {"type": "string", "example": "北京市"},
                                "children": {"type": "array"}
                            }
                        }
                    }
                }
            }
        }
    )
    @action(detail=False, methods=['get'], url_path='tree')
    def get_tree(self, request):
        """获取树形结构数据"""
        depth = int(request.query_params.get('depth', 3))

        # 获取所有省份
        provinces = Area.objects.filter(parent__isnull=True).order_by('code')
        result = []

        for province in provinces:
            province_data = {
                'value': province.code,
                'label': province.name,
            }

            if depth >= 2:
                # 获取城市
                cities = Area.objects.filter(parent=province).order_by('code')
                city_list = []

                for city in cities:
                    city_data = {
                        'value': city.code,
                        'label': city.name,
                    }

                    if depth >= 3:
                        # 获取区县
                        districts = Area.objects.filter(parent=city).order_by('code')
                        district_list = [{
                            'value': district.code,
                            'label': district.name,
                        } for district in districts]

                        city_data['children'] = district_list

                    city_list.append(city_data)

                province_data['children'] = city_list

            result.append(province_data)

        return ApiResponse(data=result, message=_("查询成功"))

    @extend_schema(
        summary="根据代码获取区域信息",
        description="通过行政区划代码获取区域详细信息及父子关系",
        parameters=[
            OpenApiParameter(name="code", type=str, required=True,
                             location=OpenApiParameter.QUERY,
                             description="行政区划代码，如：110105（北京市朝阳区）"),
        ],
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "data": {
                        "type": "object",
                        "properties": {
                            "current": {"$ref": "#/components/schemas/Area"},
                            "parent": {"$ref": "#/components/schemas/Area"},
                            "children": {"type": "array", "items": {"$ref": "#/components/schemas/Area"}}
                        }
                    }
                }
            },
            404: {"description": "区域不存在"}
        }
    )
    @action(detail=False, methods=['get'], url_path='detail-by-code')
    def detail_by_code(self, request):
        """根据代码获取区域详情"""
        code = request.query_params.get('code')
        if not code:
            return ApiResponse(code=400, message=_("缺少区域代码参数"))
        try:
            area = Area.objects.get(code=code)
            # 序列化当前区域
            current_data = self.get_serializer(area).data
            # 获取父级区域
            parent_data = None
            if area.parent:
                parent_data = self.get_serializer(area.parent).data
            # 获取子区域
            children = area.children.all().order_by('code')
            children_data = self.get_serializer(children, many=True).data
            response_data = {
                'current': current_data,
                'parent': parent_data,
                'children': children_data
            }
            return ApiResponse(data=response_data, message=_("查询成功"))
        except Area.DoesNotExist:
            return ApiResponse(code=404, message=_("区域不存在"))
    @extend_schema(
        summary="批量导入省市区数据（Excel）",
        description="通过上传 Excel 文件批量导入省市区数据，支持.xlsx 格式",
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {
                        "type": "string",
                        "format": "binary",
                        "description": "Excel 文件（.xlsx 格式）"
                    }
                },
                "required": ["file"]
            }
        },
        responses={
            200: {
                "type": "object",
                "properties": {
                    "code": {"type": "integer", "example": 200},
                    "message": {"type": "string", "example": "导入完成！成功：3000 条，失败：0 条"},
                    "data": {
                        "type": "object",
                        "properties": {
                            "success_count": {"type": "integer", "example": 3000},
                            "fail_count": {"type": "integer", "example": 0},
                            "created_count": {"type": "integer", "example": 3000},
                            "updated_count": {"type": "integer", "example": 0},
                            "errors": {"type": "array", "items": {"type": "string"}}
                        }
                    }
                }
            },
            400: {"description": "文件格式错误或参数错误"},
            500: {"description": "服务器内部错误"}
        }
    )
    @action(detail=False, methods=['post'], url_path='batch-import')
    def batch_import(self, request):
        """
        批量导入省市区数据（Excel 文件）

        Excel 格式要求：
        - 第一行：表头（code, name, parent_code）
        - 从第二行开始：数据行
        - code: 6 位行政区划代码（必填）
        - name: 区域名称（必填）
        - parent_code: 父级区域代码（可选，省级为 null 或空）

        示例：
        code      | name      | parent_code
        110000    | 北京市     |
        110100    | 北京市辖区 | 110000
        110101    | 东城区     | 110100
        """
        from openpyxl import load_workbook

        try:
            # 1. 获取上传的文件
            file = request.FILES.get('file')
            if not file:
                return ApiResponse(code=400, message=_("请上传 Excel 文件"))

            # 2. 验证文件格式
            if not file.name.endswith('.xlsx'):
                return ApiResponse(code=400, message=_("仅支持.xlsx 格式的 Excel 文件"))

            # 3. 读取 Excel 文件
            wb = load_workbook(file, data_only=True)
            ws = wb.active

            # 4. 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).strip().lower() if cell.value else "")

            # 5. 验证必要字段
            required_columns = ['code', 'name']
            if not all(col in headers for col in required_columns):
                return ApiResponse(
                    code=400,
                    message=_("Excel 文件缺少必要列，需要包含：%(cols)s") % {"cols": ", ".join(required_columns)}
                )

            # 6. 定义数据处理函数
            def process_cell_value(value):
                """处理单元格值，转换类型和处理空值"""
                if value is None or str(value).strip().lower() in ['nan', 'n/a', '', 'none']:
                    return None
                return str(value).strip()

            # 7. 解析所有数据行
            data_list = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 将行数据转换为字典
                    row_dict = dict(zip(headers, row))

                    code = process_cell_value(row_dict.get('code'))
                    name = process_cell_value(row_dict.get('name'))
                    parent_code = process_cell_value(row_dict.get('parent_code'))

                    # 验证必要字段
                    if not code or not name:
                        raise ValueError("代码和名称不能为空")

                    # 验证代码格式（6 位数字）
                    if len(code) != 6 or not code.isdigit():
                        raise ValueError("行政区划代码必须为 6 位数字")

                    data_list.append({
                        'code': code,
                        'name': name,
                        'parent_code': parent_code
                    })

                except Exception as e:
                    logger.error(f"第{row_idx}行数据解析失败：{str(e)}")
                    continue

            if not data_list:
                return ApiResponse(code=400, message=_("Excel 文件中没有有效数据"))

            # 8. 批量导入数据
            success_count = 0
            fail_count = 0
            created_count = 0
            updated_count = 0
            error_messages = []

            # 先收集所有代码，便于后续查找父级
            all_codes = {item['code'] for item in data_list}

            # 第一次遍历：创建所有记录（不设置父子关系）
            for index, item in enumerate(data_list):
                try:
                    code = item['code']
                    name = item['name']

                    # 创建或更新区域
                    area, created = Area.objects.update_or_create(
                        code=code,
                        defaults={'name': name}
                    )

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                    success_count += 1

                except Exception as e:
                    fail_count += 1
                    error_msg = f"代码{item.get('code')}导入失败：{str(e)}"
                    error_messages.append(error_msg)
                    logger.error(error_msg)

            # 第二次遍历：设置父子关系
            for item in data_list:
                try:
                    code = item['code']
                    parent_code = item.get('parent_code')

                    if parent_code:
                        try:
                            area = Area.objects.get(code=code)
                            parent = Area.objects.get(code=parent_code)

                            if area.parent_id != parent.id:
                                area.parent = parent
                                area.save(update_fields=['parent'])

                        except Area.DoesNotExist:
                            logger.warning(f"设置父子关系失败 {code}: 父级区域{parent_code}不存在")
                        except Exception as e:
                            logger.warning(f"设置父子关系失败 {code}: {str(e)}")

                except Exception as e:
                    logger.warning(f"处理父子关系失败 {item.get('code')}: {str(e)}")

            # 9. 构造返回消息
            message = f"导入完成！成功：{success_count}条，失败：{fail_count}条，新增：{created_count}条，更新：{updated_count}条"

            result_data = {
                "success_count": success_count,
                "fail_count": fail_count,
                "created_count": created_count,
                "updated_count": updated_count,
            }

            if error_messages:
                result_data["errors"] = error_messages[:50]  # 只显示前 50 个错误

            return ApiResponse(data=result_data, message=message)

        except Exception as e:
            logger.error(f"批量导入失败：{str(e)}", exc_info=True)
            return ApiResponse(code=500, message=_("批量导入失败：%(error)s") % {"error": str(e)})

# ... existing code ...
